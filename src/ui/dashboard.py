from __future__ import annotations

import re
from datetime import datetime
import threading

import flet as ft
from src.config.theme import get_colors, ACCENT, rgba
from src.core.constants import VERSION_CHECK_URL, get_local_version, SPECIAL_WAREHOUSE_RE
from src.core.processor import (
    load_lineas,
    calcular_kpis_almacen,
    obtener_metricas_lineas,
    obtener_metricas_categorias,
    sugerir_transferencias,
    export_to_excel,
    _extract_report_skus,
    _make_report_name,
    update_catalogo_sku,
    _sku_info,
    export_catalogo_to_excel,
)
from src.core.s1_downloader import get_api_sku_meta, download_catalogo
from src.ui.warehouse_card import WarehouseCard
from src.ui.linea_section import LineaSection

try:
    from g360_flet.g360_signature import G360Signature
except ImportError:
    G360Signature = None


NUM_FONT = "JetBrains Mono"


class Dashboard:
    def __init__(self, page: ft.Page, theme_mode: str = "dark"):
        self.page = page
        self._theme_mode = theme_mode
        self.c = get_colors(self._theme_mode)
        self._search_timer = None
        self._warehouse_cards: ft.Column | None = None
        self._cat_section: ft.Container | None = None
        self._kpi_row: ft.Row | None = None
        self._header: ft.Container | None = None
        self._sidebar: ft.Container | None = None
        self._sidebar_chips: ft.Column | None = None
        self._main_container: ft.Container | None = None
        self._body_listview: ft.Container | None = None
        self._main_content_area: ft.Container | None = None
        self._sin_cat_badge_text: ft.Text | None = None
        self._sin_cat_row: ft.Container | None = None
        self._search_field = ft.TextField(
            hint_text="Buscar SKU, descripción, línea o categoría…  (Enter ↵ · Esc)",
            border_radius=8,
            height=40,
            text_size=13,
            dense=True,
            border=ft.InputBorder.OUTLINE,
            border_color=self.c["border"],
            focused_border_color=self.c["accent"],
            cursor_color=self.c["accent"],
            hint_style=ft.TextStyle(size=12, color=self.c["text_muted"]),
            suffix_icon=ft.Icon(ft.Icons.KEYBOARD_RETURN, size=16, color=rgba(self.c["text_muted"], 0.5)),
            on_change=lambda e: self._on_search_change(),
            on_submit=self._on_search_submit,
            on_focus=lambda e: self._on_search_focus(),
            expand=True,
            suffix=ft.IconButton(
                icon=ft.Icons.CLEAR,
                icon_size=16,
                icon_color=self.c["text_muted"],
                tooltip="Limpiar búsqueda",
                visible=False,
                on_click=lambda e: self._clear_search(),
            ),
        )
        self._selected_alms: set[str] = set()
        self._warehouse_group: str = "venta"  # "venta" | "mktd"
        self._transfer_section = ft.Container(visible=False)
        self._transfer_collapsed = False
        self._transfer_sort_key: str | None = None
        self._transfer_sort_rev = False
        self._sin_cat_count = 0
        self._file_picker = ft.FilePicker()
        self._active_dlg = None
        self._filtro_salud = "todo"
        self._theme_button: ft.IconButton | None = None
        self._filtro_salud_row = ft.Container()
        self._build_filtro_salud()
        # Pills de vista de almacenes
        self._alm_view_pills = ft.Container()
        self._build_alm_view_pills()
        self._card_instances: list = []
        self._ts_text = ft.Text(
            "", size=11, color=self.c["text_muted"], weight=ft.FontWeight.W_500,
        )
        self._stale_badge = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.WARNING_AMBER, size=14, color=self.c["warning"]),
                ft.Text("Datos en caché", size=11, color=self.c["warning"], weight=ft.FontWeight.W_500),
            ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            visible=False,
            padding=ft.Padding(left=10, right=10, top=5, bottom=5),
            bgcolor=rgba(self.c["warning"], 0.07),
            border_radius=8,
            tooltip="Los datos son en caché porque están fuera del horario de actualización (domingo o fuera de 7:00-22:59)",
        )
        self.status_text = ft.Text(
            "Cargue datos de stock para comenzar",
            size=13, color=self.c["text_muted"], weight=ft.FontWeight.W_500,
        )
        self._raw_data: dict | None = None
        self._kpis_alm: dict | None = None
        self._lineas: list | None = None
        self._categorias: list | None = None
        self._cache_timestamp: str | None = None
        self._api_timestamp: str | None = None
        self._stale_data = False
        self._health_badge: ft.Container | None = None
        self._refresh_status_badge: ft.Container | None = None
        self._refreshing: bool = False

    def build(self) -> ft.Container:
        self._init_loading()
        self._header = self._build_header()
        self._kpi_row = self._build_kpi_row({})
        self._warehouse_cards = ft.Column(spacing=10)
        self._cat_section = ft.Container(visible=False)
        self._body_divider = ft.Divider(height=1, color=self.c["border"])
        self._empty_state = ft.Container(
            visible=False,
            content=ft.Column([
                ft.Icon(ft.Icons.CLOUD_OFF, size=48, color=rgba(self.c["text_muted"], 0.35)),
                ft.Text("Sin datos disponibles", size=16, color=self.c["text_muted"], weight=ft.FontWeight.W_600),
                ft.Text("Verificando conexión con el servidor...", size=13, color=rgba(self.c["text_muted"], 0.65)),
                ft.Container(
                    content=ft.Text("", size=11, color="white", weight=ft.FontWeight.W_600),
                    bgcolor=self.c["text_muted"],
                    border_radius=12,
                    padding=ft.Padding(left=12, right=12, top=4, bottom=4),
                    visible=False,
                ),
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=12),
        )

        self._body_listview = ft.Container(
            content=ft.Column([
                self._warehouse_cards,
                self._body_divider,
                self._cat_section,
                self._empty_state,
            ], spacing=8),
            padding=ft.Padding(left=20, right=20, top=0, bottom=0),
        )
        self._body_listview.bgcolor = self.c["background"]
        
        self._signature = G360Signature(mode="powered", version=get_local_version()) if G360Signature \
            else ft.Text("Powered by G360", size=10, color=rgba(self.c["accent"], 0.45), weight=ft.FontWeight.W_600)

        self._main_content_area = ft.Container(
            content=ft.Stack([
                ft.Container(
                    content=ft.Column([
                        self._header,
                        self._loading_bar,
                        self._kpi_row,
                        ft.Column([
                            self._transfer_section,
                            self._body_listview,
                            ft.Container(
                                content=self.status_text,
                                padding=ft.Padding(left=20, right=20, top=6, bottom=6),
                            ),
                        ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True),
                    ], spacing=0),
                    expand=True,
                ),
                ft.Container(
                    content=self._signature,
                    right=16,
                    bottom=8,
                ),
            ], expand=True),
            expand=True,
            bgcolor=self.c["background"],
        )
        self._sidebar = self._build_sidebar()
        self._main_container = ft.Container(
            content=ft.Row([self._sidebar, self._main_content_area], spacing=0),
            expand=True,
            bgcolor=self.c["background"],
        )
        # Los FilePicker deben agregarse al overlay de la página para poder usarse
        # (save_file/pick_files lanzan AssertionError si no están registrados).

        return self._main_container

    def register_overlay(self):
        self.page.overlay.extend([self._file_picker])
        self.page.update()

    def _build_sidebar(self) -> ft.Container:
        self._sin_cat_badge_text = ft.Text("", size=10, color="white", weight=ft.FontWeight.W_700)
        self._sin_cat_badge = ft.Container(
            content=self._sin_cat_badge_text,
            bgcolor=self.c["warning"], border_radius=8,
            padding=ft.Padding(left=6, right=6, top=2, bottom=2),
            visible=False,
        )
        self._sidebar_chips = ft.Column(spacing=3, scroll=ft.ScrollMode.AUTO, expand=True)
        self._sin_cat_row = ft.Container(
            content=ft.GestureDetector(
                content=ft.Row([
                    ft.Icon(ft.Icons.HELP_OUTLINE, size=14, color=self.c["warning"]),
                    ft.Text("SKUs sin categoría", size=10, color=self.c["text_muted"]),
                    ft.Container(expand=True),
                    self._sin_cat_badge,
                ], spacing=4),
                on_tap=lambda e: self._show_sin_linea_dlg(),
                mouse_cursor=ft.MouseCursor.CLICK,
            ),
            padding=ft.Padding(left=12, right=12, top=8, bottom=8),
            visible=False,
        )
        self._theme_button = ft.IconButton(
            icon=ft.Icons.DARK_MODE if self._theme_mode == "dark" else ft.Icons.LIGHT_MODE,
            icon_size=16,
            icon_color=self.c["violet"] if self._theme_mode == "dark" else self.c["orange"],
            on_click=lambda _: self._on_theme_toggle_click(),
            tooltip="Tema claro/oscuro",
        )
        self._config_button = ft.IconButton(
            icon=ft.Icons.SETTINGS,
            icon_size=16,
            icon_color=self.c["accent"],
            on_click=self._open_config,
            tooltip="Configurar almacenes",
            style=ft.ButtonStyle(bgcolor={"hovered": rgba(self.c["accent"], 0.14)}),
        )
        self._footer_chip = ft.Container(
                content=ft.Row([
                    self._theme_button,
                    ft.VerticalDivider(width=1, color=self.c["border"]),
                    self._config_button,
                ], spacing=2, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                border_radius=8,
                padding=ft.Padding(left=4, right=4, top=3, bottom=3),
        )

        # Construcción limpia del sidebar sin referencias a los pickers en los controls
        return ft.Container(
            content=ft.Stack([
                ft.Column([
                    ft.Container(
                        content=ft.Row([
                            self._search_field,
                            ft.IconButton(
                                icon=ft.Icons.SEARCH,
                                icon_size=18,
                                icon_color=self.c["text_muted"],
                                tooltip="Buscar (Enter)",
                                on_click=self._on_search_submit,
                            ),
                        ], spacing=4),
                        padding=ft.Padding(left=8, right=8, top=14, bottom=8),
                    ),
                    ft.Container(
                        content=ft.Row([
                            ft.Text("ALMACENES", size=10, color=self.c["text_muted"], weight=ft.FontWeight.W_700),
                            ft.Container(expand=True),
                            self._alm_view_pills,
                        ], spacing=4),
                        padding=ft.Padding(left=12, right=12, top=6, bottom=4),
                    ),
                    self._sidebar_chips,
                    ft.Divider(height=1, color=self.c["border"]),
                    self._sin_cat_row,
                    self._filtro_salud_row,
                    ft.Container(expand=True),
                    ft.Divider(height=1, color=self.c["border"]),
                    ft.Container(
                        content=ft.Row([
                            self._footer_chip,
                            ft.Container(expand=True),
                        ], spacing=0),
                        padding=ft.Padding(left=8, right=8, top=6, bottom=6),
                    ),
                ], spacing=0),
            ], expand=True),
            width=200,
            bgcolor=self.c["surface"],
            border=ft.Border(right=ft.BorderSide(1, self.c["border"])),
        )

    def _init_loading(self):
        self._loading_bar = ft.Container(
            visible=False,
            content=ft.Row([
                ft.ProgressRing(width=22, height=22, stroke_width=3, color="white"),
                ft.Text("Procesando...", size=14, weight=ft.FontWeight.W_700, color="white"),
            ], spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor=self.c["accent_dark"],
            border_radius=0,
            padding=ft.Padding(left=20, right=20, top=10, bottom=10),
            animate_opacity=300,
        )

    def set_loading(self, active: bool, msg: str = ""):
        self._loading_bar.visible = active
        if msg and self._loading_bar.content and len(self._loading_bar.content.controls) > 1:
            self._loading_bar.content.controls[1].value = msg
        self.page.update()

    def _build_header(self):
        logo = ft.Image(
            src="assets/images/Logo_cipsa_solid.svg",
            width=32, height=32,
        )
        self._ts_badge = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.SCHEDULE, size=14, color=rgba(self.c["accent"], 0.75)),
                self._ts_text,
            ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            visible=False,
            padding=ft.Padding(left=10, right=10, top=5, bottom=5),
            bgcolor=rgba(self.c["accent"], 0.07),
            border_radius=8,
        )
        self._health_badge = ft.Container(
            content=ft.Row([
                ft.Text("", size=11, weight=ft.FontWeight.W_600, font_family=NUM_FONT),
            ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            visible=False,
            padding=ft.Padding(left=10, right=10, top=5, bottom=5),
            border_radius=8,
            animate_opacity=200,
        )
        self._refresh_status_badge = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.CIRCLE, size=10, color=self.c["text_muted"]),
                ft.Text("—", size=11, weight=ft.FontWeight.W_500, font_family=NUM_FONT, color=self.c["text_muted"]),
            ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            visible=False,
            padding=ft.Padding(left=10, right=10, top=5, bottom=5),
            border_radius=8,
            ink=True,
            on_click=self._on_refresh,
            tooltip="Estado de la data: autoactualizable cada 15 min. Click para forzar actualización.",
        )
        self._update_banner = ft.Container(
            visible=False,
            data="update_banner",
            content=ft.Row([
                ft.Icon(ft.Icons.SYSTEM_UPDATE, size=14, color="white"),
                ft.Text("Nueva versión", size=12, color="white", weight=ft.FontWeight.W_600),
                ft.Text("", size=12, color="white", data="version_label"),
                ft.TextButton(
                    "Actualizar",
                    on_click=lambda e: self._open_update_url(),
                    style=ft.ButtonStyle(padding=ft.Padding(left=8, right=8, top=4, bottom=4), color="white"),
                ),
            ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor=self.c["warning"],
            border_radius=8,
            padding=ft.Padding(left=12, right=12, top=6, bottom=6),
        )
        self._offline_badge = ft.Container(
            visible=False,
            content=ft.Row([
                ft.Icon(ft.Icons.WIFI_OFF, size=14, color="white"),
                ft.Text("Sin conexión", size=11, color="white", weight=ft.FontWeight.W_600),
            ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor=self.c["error"],
            border_radius=8,
            padding=ft.Padding(left=10, right=10, top=5, bottom=5),
            animate_opacity=200,
        )
        return ft.Container(
            content=ft.Row([
                logo,
                ft.Text("Stock Monitor", size=20, weight=ft.FontWeight.W_800, color=self.c["accent"]),
                ft.Text("CIPSA", size=11, color=self.c["text_muted"], weight=ft.FontWeight.W_300),
                ft.Container(expand=True),
                self._update_banner,
                self._ts_badge,
                self._stale_badge,
                self._offline_badge,
                self._refresh_status_badge,
                ft.ElevatedButton(
                    "Actualizar",
                    icon=ft.Icons.REFRESH,
                    style=ft.ButtonStyle(
                        color={"": "#ffffff"},
                        bgcolor={"": self.c["accent_dark"], "hovered": self.c["accent"]},
                        shape=ft.RoundedRectangleBorder(radius=10),
                        padding=ft.Padding(left=18, right=18, top=11, bottom=11),
                    ),
                    on_click=self._on_refresh,
                ),
            ], spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            padding=ft.Padding(left=24, right=24, top=12, bottom=12),
            bgcolor=self.c["surface"],
            border=ft.Border(bottom=ft.BorderSide(1, self.c["border"])),
        )

    async def _on_refresh(self, e):
        self._set_refreshing(True)
        try:
            if self._on_refresh_cb:
                cb = self._on_refresh_cb()
                if hasattr(cb, '__await__'):
                    await cb
                else:
                    cb()
        finally:
            self._set_refreshing(False)
            self.page.update()

    def set_on_theme_toggle(self, callback):
        self._on_theme_toggle_cb = callback

    def _on_theme_toggle_click(self):
        if hasattr(self, "_on_theme_toggle_cb"):
            self._on_theme_toggle_cb()

    def update_theme(self, theme_mode: str):
        self._theme_mode = theme_mode
        self.c = get_colors(theme_mode)

        # Actualizar fondos principales
        if self._main_container:
            self._main_container.bgcolor = self.c["background"]
        if self._main_content_area:
            self._main_content_area.bgcolor = self.c["background"]
        if self._body_listview:
            self._body_listview.bgcolor = self.c["background"]

        # Actualizar Sidebar
        self._sidebar.bgcolor = self.c["surface"]
        self._sidebar.border = ft.Border(right=ft.BorderSide(1, self.c["border"]))

        # Actualizar Header
        self._header.bgcolor = self.c["surface"]
        self._header.border = ft.Border(bottom=ft.BorderSide(1, self.c["border"]))

        # Actualizar fila de KPIs
        if self._kpi_row:
            self._kpi_row.bgcolor = self.c["surface"]
            self._kpi_row.border = ft.Border(bottom=ft.BorderSide(1, self.c["border"]))

        # Actualizar componentes de entrada
        self._search_field.border_color = self.c["border"]
        self._search_field.hint_style = ft.TextStyle(size=13, color=self.c["text_muted"])

        # Actualizar iconos
        self._update_theme_button_icon()
        if self._config_button:
            self._config_button.icon_color = self.c["accent"]
        if self._footer_chip:
            self._footer_chip.bgcolor = rgba(self.c["surface_variant"], 0.7)
            if self._footer_chip.content and len(self._footer_chip.content.controls) > 1:
                self._footer_chip.content.controls[1].color = self.c["border"]
        if self._signature:
            is_mounted = getattr(self._signature, "_is_mounted", None)
            if is_mounted and is_mounted():
                self._signature.update()

        # Actualizar texto de carga
        if self._loading_bar and self._loading_bar.content:
            self._loading_bar.content.controls[1].color = self.c["text_primary"]

        # Actualizar colores de tema sin reconstruir datos
        try:
            self._refresh_theme_colors()
        except Exception as _e:
            import traceback
            traceback.print_exc()
            self.page.update()

    def _update_theme_button_icon(self):
        if self._theme_button:
            self._theme_button.icon = ft.Icons.LIGHT_MODE if self._theme_mode == "light" else ft.Icons.DARK_MODE
            self._theme_button.icon_color = self.c["violet"] if self._theme_mode == "dark" else self.c["orange"]

    def set_update_available(self, update_info: dict | None):
        if not hasattr(self, "_update_banner") or not self._update_banner:
            return
        if update_info:
            self._update_banner.visible = True
            remote_version = update_info.get("remote_version", "")
            for ctrl in self._update_banner.content.controls:
                if isinstance(ctrl, ft.Text) and ctrl.data == "version_label":
                    ctrl.value = f" v{remote_version} disponible"
        else:
            self._update_banner.visible = False
        if self.page:
            self.page.update()

    def _open_update_url(self):
        url = None
        if hasattr(self, "_update_info") and isinstance(self._update_info, dict):
            url = self._update_info.get("url")
        if not url:
            url = f"{VERSION_CHECK_URL}/../releases/latest"
        import webbrowser
        webbrowser.open(url)

    def set_on_refresh(self, callback):
        self._on_refresh_cb = callback


    def _open_config(self, e):
        self._show_config_dialog()

    def _build_config_warehouse_row(self, cod: str, cfg: dict) -> tuple[ft.Container, ft.TextField, ft.Dropdown, ft.Dropdown]:
        prio = ft.TextField(
            value=str(cfg.get("prioridad", "")),
            width=50, height=36, text_size=12,
            text_align=ft.TextAlign.CENTER, dense=True,
        )
        rol_dropdown = ft.Dropdown(
            value=cfg.get("rol", "EXTERNO"),
            options=[
                ft.dropdown.Option("PRINCIPAL"),
                ft.dropdown.Option("SECUNDARIO"),
                ft.dropdown.Option("EXTERNO"),
            ],
            width=110, text_size=11,
            dense=True,
            border_radius=8,
        )
        # Dropdown para tipo de almacén (venta/mktd)
        tipo_actual = cfg.get("tipo", "venta")
        tipo_dropdown = ft.Dropdown(
            value=tipo_actual,
            options=[
                ft.dropdown.Option("venta", "VENTA"),
                ft.dropdown.Option("mktd", "MKTD"),
            ],
            width=90, text_size=11,
            dense=True,
            border_radius=8,
        )
        tipo_color = {"DESAGREGADO": self.c["accent"], "CONSOLIDADO": self.c["info"], "PCT": self.c["violet"]}
        row = ft.Container(
            content=ft.Row([
                ft.Text(cod, size=13, weight=ft.FontWeight.W_700, width=50, color=self.c["text_primary"]),
                ft.Text(cfg.get("nombre", ""), size=11, color=self.c["text_muted"], expand=True),
                ft.Container(
                    content=ft.Text(cfg.get("tipo_reporte", ""), size=10, color="white", weight=ft.FontWeight.W_600),
                    bgcolor=tipo_color.get(cfg.get("tipo_reporte", ""), "#666"),
                    border_radius=5, padding=ft.Padding(left=6, right=6, top=2, bottom=2),
                    width=85, alignment=ft.alignment.center
                ),
                rol_dropdown,
                tipo_dropdown,
                prio,
            ], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            padding=ft.Padding(left=4, right=4, top=6, bottom=6),
            border=ft.Border(bottom=ft.BorderSide(1, self.c["border"]))
        )
        return row, prio, rol_dropdown, tipo_dropdown

    def _show_config_dialog(self):
        config = load_lineas()
        alm_config = config.get("almacenes", {})
        rows = []
        prio_entries = {}
        rol_entries = {}
        tipo_entries = {}

        sort_order = {"PRINCIPAL": 0, "SECUNDARIO": 1, "EXTERNO": 2}
        sorted_codes = sorted(alm_config.keys(), key=lambda c: (sort_order.get(alm_config[c].get("rol", ""), 9), alm_config[c].get("prioridad", 99)))

        rows.append(ft.Container(
            content=ft.Row([
                ft.Text("Cód", size=11, weight=ft.FontWeight.W_700, color=self.c["accent"], width=50),
                ft.Text("Nombre Almacén", size=11, weight=ft.FontWeight.W_700, color=self.c["accent"], expand=True),
                ft.Text("Reporte", size=11, weight=ft.FontWeight.W_700, color=self.c["accent"], width=85),
                ft.Text("Rol Operativo", size=11, weight=ft.FontWeight.W_700, color=self.c["accent"], width=110),
                ft.Text("Tipo", size=11, weight=ft.FontWeight.W_700, color=self.c["accent"], width=90),
                ft.Text("Prio", size=11, weight=ft.FontWeight.W_700, color=self.c["accent"], width=50),
            ], spacing=8),
            bgcolor=rgba(self.c["accent"], 0.08), border_radius=6,
            padding=ft.Padding(left=8, right=8, top=6, bottom=6),
        ))

        for cod in sorted_codes:
            cfg = alm_config[cod]
            row, prio, rol_dropdown, tipo_dropdown = self._build_config_warehouse_row(cod, cfg)
            prio_entries[cod] = prio
            rol_entries[cod] = rol_dropdown
            tipo_entries[cod] = tipo_dropdown
            rows.append(row)

        def save(e):
            for cod in alm_config.keys():
                try:
                    config["almacenes"][cod]["prioridad"] = int(prio_entries[cod].value.strip())
                    config["almacenes"][cod]["rol"] = rol_entries[cod].value
                    config["almacenes"][cod]["tipo"] = tipo_entries[cod].value
                except (ValueError, AttributeError, KeyError):
                    pass
            from src.core.constants import LINEAS_FILE
            import json
            with open(LINEAS_FILE, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            self.page.close(dlg)
            self._apply_filters() # Refrescar la vista principal con nuevos roles

        def _on_catalog_export(e):
            def on_result(pe: ft.FilePickerResultEvent):
                if not pe.path:
                    return
                try:
                    items = download_catalogo()
                    if not items:
                        raise ValueError("No se pudo descargar el catálogo del API")
                    export_catalogo_to_excel(items, pe.path)
                    self._show_snack(f"Catálogo exportado: {len(items)} SKUs")
                    import os
                    os.startfile(pe.path)
                except Exception as ex:
                    self._show_snack(f"Error al exportar catálogo: {ex}", is_error=True)

            self._file_picker.on_result = on_result
            self._file_picker.save_file(file_name=f"{_make_report_name('catálogo maestro')}.xlsx")

        catalog_export_btn = ft.ElevatedButton(
            "Exportar Catálogo XLSX",
            icon=ft.Icons.DOWNLOAD,
            on_click=_on_catalog_export,
            style=ft.ButtonStyle(bgcolor={"": self.c["surface_variant"]}, color={"": self.c["text_primary"]}),
        )

        dlg = ft.AlertDialog(
            title=ft.Row([ft.Icon(ft.Icons.TUNE, color=self.c["accent"]), ft.Text("Configuración de Almacenes", weight=ft.FontWeight.W_800, size=16)]),
            content=ft.Container(
                content=ft.Column([
                    ft.Column(rows, spacing=0, scroll=ft.ScrollMode.AUTO, expand=True),
                    ft.Divider(height=12, color=rgba(self.c["border"], 0.5)),
                    ft.Row([
                        ft.Text("Respaldo del catálogo maestro (sin stock)", size=11, color=self.c["text_muted"]),
                        catalog_export_btn,
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ], spacing=8),
                width=680, height=480
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda e: self.page.close(dlg)), 
                ft.ElevatedButton("Guardar Cambios", on_click=save, bgcolor=self.c["accent_dark"], color="white",
                                  style=ft.ButtonStyle(overlay_color=rgba(self.c["accent"], 0.1), elevation=0)),
            ],
        )
        self.page.open(dlg)

    def _sidebar_chip(self, cod: str, rol: str, selected: bool, is_special: bool = False, alertas: int = 0, criticos: int = 0, is_mktd: bool = False) -> ft.Container:
        rol_color = {"PRINCIPAL": self.c["accent"], "SECUNDARIO": self.c["info"], "EXTERNO": self.c["text_muted"]}
        base = rol_color.get(rol, "#666")
        extra = []
        if is_special:
            extra.append(ft.Icon(ft.Icons.INFO_OUTLINED, size=10, color=rgba(self.c["text_muted"], 0.5), tooltip="Almacén informativo (MKTD/S)"))
        control_badge = None
        if rol in ("PRINCIPAL", "SECUNDARIO") and (criticos > 0 or alertas > 0):
            badge_text = ""
            badge_color = self.c["text_muted"]
            if criticos > 0:
                badge_text = f"{criticos}!"
                badge_color = self.c["error"]
            elif alertas > 0:
                badge_text = f"{alertas}⚠"
                badge_color = self.c["warning"]
            control_badge = ft.Container(
                content=ft.Text(badge_text, size=9, color="white", weight=ft.FontWeight.W_700),
                bgcolor=badge_color, border_radius=6,
                padding=ft.Padding(left=4, right=4, top=2, bottom=2),
            )
        sel_border = ft.BorderSide(2, base) if selected else ft.BorderSide(0, "transparent")
        inner = ft.Container(
            content=ft.Row([
                ft.Container(width=8, height=8, bgcolor=base if selected else rgba(base, 0.6), border_radius=4),
                ft.Text(cod, size=12, weight=ft.FontWeight.W_600 if selected else ft.FontWeight.W_400, color=self.c["text_primary"] if selected else rgba(self.c["text_primary"], 0.5)),
                ft.Container(expand=True),
                *extra,
                control_badge if control_badge else ft.Container(),
                ft.Text(rol[:4], size=10, color=rgba(base, 0.6)),
            ], spacing=4),
            border_radius=6,
            bgcolor=rgba(base, 0.2) if selected else rgba(base, 0.05),
            border=ft.Border(left=sel_border),
            padding=ft.Padding(left=8, right=10, top=5, bottom=5),
        )
        return ft.GestureDetector(
            content=inner,
            on_tap=lambda e, c=cod: self._on_chip_toggle(c),
            mouse_cursor=ft.MouseCursor.CLICK,
        )

    def _build_filtro_salud(self):
        opts = [("todo", "Todo"), ("ok", "OK"), ("alerta", "Alerta"), ("critico", "Crítico")]
        pills = []
        for val, label in opts:
            is_active = val == self._filtro_salud
            pill = ft.GestureDetector(
                content=ft.Container(
                    content=ft.Text(label, size=10, weight=ft.FontWeight.W_700 if is_active else ft.FontWeight.W_500,
                                    color="white" if is_active else self.c["text_muted"]),
                    padding=ft.Padding(left=8, right=8, top=4, bottom=4),
                    bgcolor=self.c["accent"] if is_active else rgba(self.c["text_muted"], 0.08),
                    border_radius=12,
                ),
                on_tap=lambda e, v=val: self._on_filtro_salud(v),
                mouse_cursor=ft.MouseCursor.CLICK,
            )
            pills.append(pill)
        self._filtro_salud_row.content = ft.Column([
            ft.Text("Filtrar salud", size=9, weight=ft.FontWeight.W_600, color=self.c["text_muted"]),
            ft.Row(pills, spacing=4, wrap=True),
        ], spacing=4)

    def _on_filtro_salud(self, val: str):
        self._filtro_salud = val
        self._build_filtro_salud()
        self._apply_filters()

    def _is_venta(self, cod: str, alm_config: dict, raw_data: dict | None = None) -> bool:
        """Un almacén es VENTA si su tipo_config dice 'venta' (configuración manual)."""
        tipo_config = alm_config.get(cod, {}).get("tipo", "")
        if tipo_config:
            return tipo_config == "venta"
        # Fallback a datos del API
        raw = raw_data or self._raw_data or {}
        if cod in raw:
            primer_sku = next(iter(raw[cod].values()), None)
            if primer_sku:
                cat = primer_sku.get("almacen_categoria", "venta")
                return cat == "venta"
        return True

    def _is_mktd(self, cod: str, alm_config: dict, raw_data: dict | None = None) -> bool:
        """Un almacén es MKTD si su tipo_config dice 'mktd' o es un s* sin datos."""
        tipo_config = alm_config.get(cod, {}).get("tipo", "")
        if tipo_config:
            return tipo_config == "mktd"
        raw = raw_data or self._raw_data or {}
        if cod in raw:
            primer_sku = next(iter(raw[cod].values()), None)
            if primer_sku:
                cat = primer_sku.get("almacen_categoria", "mktd")
                return cat == "mktd"
        return bool(re.match(r"^s\d+$", cod, re.IGNORECASE))

    def _build_alm_view_pills(self):
        """Botón toggle VENTA / MKTD."""
        label = "MKTD" if self._warehouse_group == "mktd" else "VENTA"
        color = self.c["info"] if self._warehouse_group == "mktd" else self.c["accent"]
        self._alm_view_pills.content = ft.GestureDetector(
            content=ft.Text(label, size=9, weight=ft.FontWeight.W_700, color=color),
            on_tap=lambda e: self._on_warehouse_group_change("mktd" if self._warehouse_group == "venta" else "venta"),
            mouse_cursor=ft.MouseCursor.CLICK,
        )

    def _on_warehouse_group_change(self, group: str):
        """Cambia entre 'venta' y 'mktd'. Mutuamente excluyente."""
        self._warehouse_group = group
        # Actualizar pills
        self._build_alm_view_pills()
        if self._alm_view_pills.content:
            try:
                self._alm_view_pills.content.update()
            except Exception:
                pass
        raw = self._raw_data or {}
        if not raw:
            return
        alm_config = self._get_alm_config(raw)
        if group == "venta":
            self._selected_alms = {c for c in raw.keys() if self._is_venta(c, alm_config, raw)}
        else:
            self._selected_alms = {c for c in raw.keys() if self._is_mktd(c, alm_config, raw)}
        self._apply_filters()

    def _on_chip_toggle(self, cod: str):
        raw = self._raw_data or {}
        alm_config = self._get_alm_config(raw)
        is_mktd_chip = self._is_mktd(cod, alm_config, raw)
        if self._warehouse_group == "mktd" and is_mktd_chip:
            # MKTD mode: toggle todo el grupo
            mktd_codes = {c for c in raw.keys() if self._is_mktd(c, alm_config, raw)}
            if mktd_codes.issubset(self._selected_alms):
                self._selected_alms -= mktd_codes
            else:
                self._selected_alms |= mktd_codes
        elif self._warehouse_group == "venta" and not is_mktd_chip:
            # VENTA mode: toggle individual
            if cod in self._selected_alms:
                self._selected_alms.discard(cod)
            else:
                self._selected_alms.add(cod)
        elif is_mktd_chip:
            # En modo VENTA, clickear un chip MKTD → cambiar a modo MKTD
            self._warehouse_group = "mktd"
            mktd_codes = {c for c in raw.keys() if self._is_mktd(c, alm_config, raw)}
            self._selected_alms = mktd_codes.copy()
            self._build_alm_view_pills()
            if self._alm_view_pills.content:
                try:
                    self._alm_view_pills.content.update()
                except Exception:
                    pass
        else:
            # En modo MKTD, clickear un chip VENTA → cambiar a modo VENTA
            self._warehouse_group = "venta"
            venta_codes = {c for c in raw.keys() if self._is_venta(c, alm_config, raw)}
            self._selected_alms = venta_codes.copy()
            self._build_alm_view_pills()
            if self._alm_view_pills.content:
                try:
                    self._alm_view_pills.content.update()
                except Exception:
                    pass
        self._apply_filters()

    def _on_search_change(self):
        """Actualiza el botón 'limpiar' y dispara búsqueda en vivo."""
        value = (self._search_field.value or "").strip()
        if hasattr(self._search_field, 'suffix') and self._search_field.suffix:
            self._search_field.suffix.visible = bool(value)
            try:
                self._search_field.suffix.update()
            except Exception:
                pass
        self._on_search()

    def _on_search_focus(self, e=None):
        """Al enfocar la barra de búsqueda, selecciona el texto para facilitar la edición."""
        try:
            if self._search_field.value:
                self._search_field.update()
        except Exception:
            pass

    def _on_search(self, immediate: bool = False):
        """Ejecuta la búsqueda filtrando el dashboard."""
        if self._search_timer:
            self._search_timer.cancel()
            self._search_timer = None
        if immediate:
            self._apply_filters()
            return
        self._search_timer = threading.Timer(0.3, self._apply_filters)
        self._search_timer.daemon = True
        self._search_timer.start()

    def _on_search_submit(self, e):
        value = (self._search_field.value or "").strip()
        if not value:
            return
        sku = value.replace("'", "").strip().upper()
        skus = self._get_all_skus()
        if sku in skus:
            self._show_sku_detail_modal(sku)
        else:
            self._on_search(immediate=True)

    def _get_all_skus(self) -> set:
        skus = set()
        if self._raw_data:
            for alm, skus_dict in self._raw_data.items():
                if alm in self._selected_alms:
                    skus.update(skus_dict.keys())
        return skus

    def _show_search_modal(self):
        """Modal de busqueda avanzada — filtra sin rebuild del dashboard completo."""
        if not self._raw_data:
            return
        raw = self._get_filtered_raw()
        all_items: list[tuple[str, str, str, int, int, int, dict]] = []
        for alm, skus_dict in raw.items():
            for sku, info in skus_dict.items():
                stock = info.get("stock", 0)
                pred = info.get("predespacho", 0)
                disp = info.get("disponible", max(0, stock - pred))
                desc = info.get("descripcion", "")[:50]
                st = self._sku_state(sku, disp)
                all_items.append((sku, desc, alm, stock, pred, disp, st))

        search_field = ft.TextField(
            hint_text="Buscar SKU, descripción...",
            prefix_icon=ft.Icons.SEARCH,
            border_radius=8,
            height=40,
            text_size=13,
            dense=True,
            border=ft.InputBorder.OUTLINE,
            border_color=self.c["border"],
            focused_border_color=self.c["accent"],
            cursor_color=self.c["accent"],
            hint_style=ft.TextStyle(size=13, color=self.c["text_muted"]),
            expand=True,
        )

        results_col = ft.Column(spacing=1, scroll=ft.ScrollMode.AUTO, expand=True)
        status_text = ft.Text("", size=10, color=self.c["text_muted"], text_align=ft.TextAlign.LEFT)

        dlg_holder = {}

        def close_dlg():
            if "dlg" in dlg_holder:
                try:
                    self.page.close(dlg_holder["dlg"])
                except Exception:
                    pass

        def do_search(e=None):
            q = (search_field.value or "").strip().lower()
            results_col.controls.clear()
            if not q:
                status_text.value = ""
                try:
                    results_col.update()
                    status_text.update()
                except Exception:
                    pass
                return
            filtered = [
                (sku, desc, alm, stock, pred, disp, st,
                 sku.upper().replace("'", "").strip() == q.upper().replace("'", "").strip())
                for sku, desc, alm, stock, pred, disp, st in all_items
                if q in sku.lower() or q in desc.lower() or q in alm.lower()
            ]
            total = len(filtered)
            MAX_ROWS = 200
            shown = filtered[:MAX_ROWS]
            for sku, desc, alm, stock, pred, disp, st, is_exact in shown:
                results_col.controls.append(ft.Container(
                    content=ft.Row([
                        ft.Text(sku, size=11, weight=ft.FontWeight.W_700, width=80, font_family=NUM_FONT),
                        ft.Text(desc, size=11, expand=True, max_lines=1, overflow=ft.TextOverflow.ELLIPSIS),
                        ft.Text(alm, size=10, width=45, color=self.c["text_muted"]),
                        ft.Text(f"s:{stock:,}", size=10, width=55, color=self.c["text_muted"], font_family=NUM_FONT, text_align=ft.TextAlign.RIGHT),
                        ft.Text(f"d:{disp:,}", size=10, weight=ft.FontWeight.W_700, width=55,
                                color=st["color"], font_family=NUM_FONT, text_align=ft.TextAlign.RIGHT),
                        ft.Container(
                            content=ft.Text(st["emoji"], size=10),
                            tooltip=f"Salud: {st['nivel']}",
                            width=16, height=16,
                        ),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    padding=ft.Padding(left=8, right=8, top=4, bottom=4),
                    bgcolor=self.c["surface_sunken"] if is_exact else None,
                    border=ft.Border(bottom=ft.BorderSide(1, self.c["border"])),
                    on_click=lambda e, s=sku, ex=is_exact: self._apply_modal_search(s, ex, close_dlg),
                    ink=True,
                ))
            if total == 0:
                status_text.value = "Sin resultados"
                status_text.color = self.c["error"]
            else:
                status_text.value = f"{total} resultado(s)" + (" (mostrando primeros 200)" if total > MAX_ROWS else "")
                status_text.color = self.c["text_muted"]
            try:
                results_col.update()
                status_text.update()
            except Exception:
                pass

        def on_submit(e):
            q = (search_field.value or "").strip()
            if not q:
                return
            sku = q.upper().replace("'", "").strip()
            if sku in {x[0] for x in all_items}:
                close_dlg()
                self._show_sku_detail_modal(sku)
            else:
                self._apply_modal_search(q, False, close_dlg)

        search_field.on_change = lambda e: do_search(e)
        search_field.on_submit = on_submit
        search_field.focus()

        def close_handler(e):
            close_dlg()

        dlg = ft.AlertDialog(
            title=ft.Text("Buscar producto", size=16, weight=ft.FontWeight.W_700),
            content=ft.Container(
                content=ft.Column([
                    ft.Row([search_field], height=40),
                    status_text,
                    results_col,
                ], spacing=4),
                width=520,
                height=380,
            ),
            actions=[
                ft.TextButton("Cerrar", on_click=close_handler),
            ],
        )
        dlg_holder["dlg"] = dlg
        self._modal_search_results = all_items
        self.page.open(dlg)
        do_search()

    def _apply_modal_search(self, value: str, exact: bool, close_fn=None):
        if close_fn:
            try:
                close_fn()
            except Exception:
                pass
        self._search_field.value = value
        if hasattr(self._search_field, 'suffix') and self._search_field.suffix:
            self._search_field.suffix.visible = bool(value)
            try:
                self._search_field.update()
            except Exception:
                pass
        if exact:
            self._on_search_submit(None)
        else:
            self._on_search(immediate=True)

    def _clear_search(self):
        if self._search_timer:
            self._search_timer.cancel()
            self._search_timer = None
        self._search_field.value = ""
        if hasattr(self._search_field, 'suffix') and self._search_field.suffix:
            self._search_field.suffix.visible = False
            try:
                self._search_field.suffix.update()
            except Exception:
                pass
        self._apply_filters()
        try:
            self._search_field.update()
        except Exception:
            pass
        self._search_field.focus()

    def _get_all_skus(self) -> set:
        raw = self._get_filtered_raw()
        alm_config = self._get_alm_config(raw)

        rows = []
        total_stock = 0
        total_disp = 0
        total_pred = 0
        cat_info = _sku_info(sku)
        desc = ""

        # Mostrar filas ordenadas por rol (PRINCIPAL primero, luego SECUNDARIO, EXTERNO)
        rol_order = {"PRINCIPAL": 0, "SECUNDARIO": 1, "EXTERNO": 2}
        sorted_almacen = sorted(
            raw.items(),
            key=lambda x: rol_order.get(alm_config.get(x[0], {}).get("rol", "EXTERNO"), 9),
        )

        for alm_cod, skus_dict in sorted_almacen:
            if sku not in skus_dict:
                continue
            info = skus_dict[sku]
            desc = info.get("descripcion", desc)
            stock = info.get("stock", 0)
            pred = info.get("predespacho", 0)
            disp = info.get("disponible", max(0, stock - pred))
            rol = alm_config.get(alm_cod, {}).get("rol", "EXTERNO")
            rol_color = {"PRINCIPAL": self.c["accent"], "SECUNDARIO": self.c["info"], "EXTERNO": self.c["text_muted"]}
            rows.append(ft.Row([
                ft.Text(alm_cod, size=12, weight=ft.FontWeight.W_600, width=50),
                ft.Text(alm_config.get(alm_cod, {}).get("nombre", alm_cod) or alm_cod, size=11, expand=True, color=self.c["text_muted"]),
                ft.Text(f"{stock:,}", size=12, width=70, color=self.c["text_primary"], font_family=NUM_FONT, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"{pred:,}", size=12, width=70, color=rgba(self.c["warning"], 0.8), font_family=NUM_FONT, text_align=ft.TextAlign.RIGHT),
                ft.Text(f"{disp:,}", size=12, weight=ft.FontWeight.W_700, width=70, color=self.c["success"], font_family=NUM_FONT, text_align=ft.TextAlign.RIGHT),
                ft.Container(
                    content=ft.Text(rol[:4], size=9, color=rol_color.get(rol, "#666")),
                    padding=ft.Padding(left=6, right=6, top=2, bottom=2),
                    bgcolor=rgba(rol_color.get(rol, "#666"), 0.1),
                    border_radius=4,
                ),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN))
            total_stock += stock
            total_disp += disp
            total_pred += pred

        st = self._sku_state(sku, total_disp)

        dlg = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.Icons.SEARCH, size=18, color=self.c["accent"]),
                ft.Text(f"SKU: {sku}", size=16, weight=ft.FontWeight.W_800),
                ft.Container(
                    content=ft.Text(st["emoji"], size=14),
                    tooltip=f"Estado: {st['nivel'].upper()}",
                ),
            ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            content=ft.Container(
                content=ft.Column([
                    ft.Text(desc or "(sin descripción)", size=13, color=self.c["text_secondary"], max_lines=2),
                    ft.Divider(height=1, color=self.c["border"]),
                    ft.Row([
                        ft.Text(f"Línea: {cat_info.get('linea_nombre') or cat_info.get('linea') or '-'}", size=11, color=self.c["text_muted"]),
                        ft.Text(f"Cat: {cat_info.get('categoria') or '-'}", size=11, color=self.c["text_muted"]),
                        ft.Text(f"U/BX: {cat_info.get('un_bx', 1)}", size=11, color=self.c["text_muted"], font_family=NUM_FONT),
                    ], spacing=12),
                    ft.Row([
                        ft.Text(f"Precio: S/ {cat_info.get('precio_lista', 0):,.2f}", size=11, color=self.c["text_muted"], font_family=NUM_FONT),
                        ft.Text(f"Min: {st['stock_minimo']} uds", size=11, color=self.c["text_muted"]),
                    ], spacing=12),
                    ft.Divider(height=1, color=self.c["border"]),
                    ft.Text("Stock por almacén", size=11, weight=ft.FontWeight.W_600, color=self.c["text_muted"]),
                    ft.Row([
                        ft.Text("Alm.", size=10, width=50, color=self.c["text_muted"], weight=ft.FontWeight.W_600),
                        ft.Text("Nombre", size=10, expand=True, color=self.c["text_muted"], weight=ft.FontWeight.W_600),
                        ft.Text("Stock", size=10, width=70, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT, weight=ft.FontWeight.W_600),
                        ft.Text("Pred.", size=10, width=70, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT, weight=ft.FontWeight.W_600),
                        ft.Text("Disp.", size=10, width=70, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT, weight=ft.FontWeight.W_600),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    ft.Container(
                        content=ft.Column(rows, spacing=4, scroll=ft.ScrollMode.AUTO),
                        height=180,
                    ),
                    ft.Divider(height=1, color=self.c["border"]),
                    ft.Row([
                        ft.Text(f"TOTAL Stock: {total_stock:,}", size=12, weight=ft.FontWeight.W_700, font_family=NUM_FONT),
                        ft.Container(expand=True),
                        ft.Text(f"Disp: {total_disp:,}", size=12, weight=ft.FontWeight.W_700, color=self.c["success"], font_family=NUM_FONT),
                    ]),
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.WARNING_AMBER, size=16, color=st["color"]),
                            ft.Text(f"Salud: {st['nivel'].upper()}", size=12, color=st["color"], weight=ft.FontWeight.W_600),
                        ], spacing=4),
                        bgcolor=rgba(st["color"], 0.08),
                        border_radius=6,
                        padding=ft.Padding(left=10, right=10, top=4, bottom=4),
                    ),
                ], spacing=8),
                width=480,
                height=400,
                padding=ft.Padding(left=16, right=16, top=16, bottom=16),
            ),
            actions=[
                ft.TextButton("Cerrar", on_click=lambda e: self.page.close(dlg)),
                ft.TextButton(
                    "Exportar",
                    icon=ft.Icons.SAVE,
                    on_click=lambda e: self._export_single_sku(sku, desc, raw),
                ),
            ],
        )
        self.page.open(dlg)

    def _export_single_sku(self, sku: str, desc: str, raw: dict):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from pathlib import Path
        wb = Workbook()
        ws = wb.active
        ws.title = sku[:31]
        headers = ["SKU", "Descripción", "Almacén", "Stock", "Predespacho", "Disponible"]
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")
        for alm_cod, skus_dict in raw.items():
            if sku not in skus_dict:
                continue
            info = skus_dict[sku]
            stock = info.get("stock", 0)
            pred = info.get("predespacho", 0)
            disp = info.get("disponible", max(0, stock - pred))
            ws.append([sku, desc or "", alm_cod, stock, pred, disp])
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        for col in "DEF":
            ws.column_dimensions[col].width = 12
        import os
        default_name = f"{_make_report_name(f'sku_{sku}')}.xlsx"
        path = str(DATA_DIR / default_name)
        wb.save(path)
        self._show_snack(f"Exportado: {default_name}")
        os.startfile(path)

    def _show_empty_state(self, message: str = "Sin datos disponibles"):
        if not self._empty_state:
            return
        self._empty_state.visible = True
        self._empty_state.content.controls[1].value = message
        if len(self._empty_state.content.controls) > 2:
            self._empty_state.content.controls[2].value = "Verificando conexión con el servidor..."
        self._warehouse_cards.visible = False
        self._body_divider.visible = False
        self._cat_section.visible = False
        self._transfer_section.visible = False
        if self._sidebar_chips:
            self._sidebar_chips.controls = []
        # No calling page.update() here - let the caller handle the final update

    def _clear_empty_state(self):
        if not self._empty_state:
            return
        self._empty_state.visible = False
        self._warehouse_cards.visible = True
        self._body_divider.visible = True
        # No calling page.update() here - let the caller handle the final update

    def _set_empty_state_status(self, text: str, color: str):
        if not self._empty_state or len(self._empty_state.content.controls) < 4:
            return
        badge = self._empty_state.content.controls[3]
        badge.content.value = text
        badge.bgcolor = color
        badge.visible = True
        self.page.update()

    def _hide_empty_state(self):
        self._clear_empty_state()

    def _get_alm_config(self, raw_data: dict | None = None) -> dict:
        """Config de almacenes enriqueciendo dinámicamente los s* (s1, s13, etc.)."""
        config = load_lineas()
        alm_config = dict(config.get("almacenes", {}))
        raw = raw_data if raw_data is not None else (self._raw_data or {})
        for cod in raw.keys():
            if SPECIAL_WAREHOUSE_RE.match(cod) and cod not in alm_config:
                # Obtener nombre del primer SKU disponible en este almacén
                primer_sku = next(iter(raw[cod].values()), None)
                nombre = ""
                if primer_sku:
                    nombre = primer_sku.get("nombre_almacen", "") or ""
                alm_config[cod] = {
                    "nombre": nombre if nombre else f"MKTD {cod}",
                    "prioridad": 9,
                    "tipo_reporte": "PCT",
                    "rol": "EXTERNO",
                    "participa_control": False,
                    "tipo": "mktd",  # s* son siempre mktd
                }
        return alm_config

    def _get_filtered_raw(self) -> dict:
        if not self._raw_data:
            return {}
        return {k: v for k, v in self._raw_data.items() if k in self._selected_alms}

    def _filter_by_health(self, raw: dict) -> dict:
        if self._filtro_salud == "todo":
            return raw
        result = {}
        for alm, skus in raw.items():
            if not isinstance(skus, dict):
                continue
            matched = {}
            for sku, info in skus.items():
                if not isinstance(info, dict):
                    continue
                try:
                    disp = info.get("disponible", info["stock"] - info["predespacho"])
                    if self._match_filtro_sku(sku, disp):
                        matched[sku] = info
                except Exception:
                    matched[sku] = info
            if matched:
                result[alm] = matched
        return result

    def _apply_filters(self):
        if not self._raw_data:
            return
        search = (self._search_field.value or "").strip().lower()
        if search:
            # Búsqueda independiente: muestra todos los almacenes
            raw = self._raw_data
        else:
            raw = self._get_filtered_raw()
        try:
            raw = self._filter_by_health(raw)
        except Exception:
            pass
        search = (self._search_field.value or "").strip().lower()
        if search:
            raw = self._apply_search_filter(raw)
            if not raw:
                self._show_no_results(search)
                return

        config = load_lineas()
        alm_config = self._get_alm_config(raw)
        self._lineas_config = {ln["codigo"]: ln for ln in config.get("lineas", [])}
        self._rebuild_kpis(raw)
        self._rebuild_warehouse_cards(raw, alm_config)
        sin_linea = self._rebuild_categoria_section(raw)
        self._update_sidebar_chips(alm_config)
        self._update_sin_linea_footer(sin_linea)
        self._update_empty_state(False)
        self._build_filtro_salud()
        self._update_health_badge()
        search_raw = raw
        transfers = sugerir_transferencias(search_raw, alm_config, umbral=5, search=search)
        self._build_transfer_section(transfers, bool(transfers), search)
        try:
            self.page.update()
        except Exception:
            pass

    def _apply_search_filter(self, raw: dict) -> dict:
        search = (self._search_field.value or "").strip().lower()
        if not search:
            return raw
        filtered = {}
        _meta_cache = {}
        for alm, skus in raw.items():
            matched = {}
            for sku, info in skus.items():
                desc = (info.get("descripcion", "") or "").lower()
                if search in sku.lower() or search in desc or search in alm.lower():
                    matched[sku] = info
                    continue
                cat_info = _meta_cache.get(sku)
                if cat_info is None:
                    cat_info = _sku_info(sku)
                    _meta_cache[sku] = cat_info
                linea = (cat_info.get("linea", "") or "").lower()
                categoria = (cat_info.get("categoria", "") or "").lower()
                if search in linea or search in categoria:
                    matched[sku] = info
            if matched:
                filtered[alm] = matched
        return filtered

    def _rebuild_kpis(self, raw: dict):
        self._kpis_alm = calcular_kpis_almacen(raw)
        if not self._kpis_alm:
            return
        total_disp = sum(a["disponible_total"] for a in self._kpis_alm.values())
        total_pre = sum(a["predespacho_total"] for a in self._kpis_alm.values())
        total_alertas = sum(a["alertas"] for a in self._kpis_alm.values())
        total_criticos = sum(a["criticos"] for a in self._kpis_alm.values())
        total_alto_stock = sum(a.get("alto_stock", 0) for a in self._kpis_alm.values())
        total_skus = sum(a["sku_count"] for a in self._kpis_alm.values())
        total_sin_catalogo = sum(a.get("sin_catalogo_count", 0) for a in self._kpis_alm.values())
        self._kpi_row.content = self._build_kpi_row({
            "almacenes": len(self._kpis_alm), "skus": total_skus,
            "disponible": total_disp, "predespacho": total_pre,
            "alertas": total_alertas, "criticos": total_criticos,
            "alto_stock": total_alto_stock, "sin_catalogo": total_sin_catalogo,
        }).content

    def _rebuild_warehouse_cards(self, raw: dict, alm_config: dict):
        if not self._kpis_alm:
            self._warehouse_cards.controls = []
            self._card_instances = []
            return
        sorted_alms = sorted(self._kpis_alm.values(), key=lambda a: alm_config.get(a["codigo"], {}).get("prioridad", 99))
        cards = []
        self._card_instances = []
        for alm in sorted_alms:
            cfg = alm_config.get(alm["codigo"], {})
            card = WarehouseCard(alm, cfg, self.c, on_click=self._show_warehouse_skus)
            container = card.build()
            card._main_container = container
            self._card_instances.append(card)
            cards.append(container)
        self._warehouse_cards.controls = cards

    def _rebuild_categoria_section(self, raw: dict):
        if not self._kpis_alm:
            return 0
        alm_config = self._get_alm_config(raw)
        self._lineas, self._lineas_sin_catalogo = obtener_metricas_lineas(self._kpis_alm, raw, alm_config)
        self._categorias, self._categorias_sin_catalogo = obtener_metricas_categorias(self._kpis_alm, raw)
        linea_section = LineaSection(self._lineas, self._categorias, self.c,
                                     on_linea_click=self._show_linea_skus,
                                     filtro_salud=self._filtro_salud,
                                     lineas_sin_catalogo=self._lineas_sin_catalogo,
                                     categorias_sin_catalogo=self._categorias_sin_catalogo)
        self._cat_section.content = linea_section.build()
        self._cat_section.visible = len(self._categorias) > 0 or len(self._categorias_sin_catalogo) > 0

    def _update_sidebar_chips(self, alm_config: dict):
        for cod, gd in self._chip_refs.items():
            cfg = alm_config.get(cod, {})
            selected = cod in self._selected_alms
            rol = cfg.get("rol", "")
            is_special = bool(SPECIAL_WAREHOUSE_RE.match(cod))
            alertas = self._kpis_alm.get(cod, {}).get("alertas", 0) if self._kpis_alm else 0
            criticos = self._kpis_alm.get(cod, {}).get("criticos", 0) if self._kpis_alm else 0
            is_mktd_chip = self._is_mktd(cod, alm_config)
            if (self._warehouse_group == "venta" and is_mktd_chip) or \
               (self._warehouse_group == "mktd" and not is_mktd_chip):
                gd.visible = False
            else:
                gd.visible = True
            new_chip = self._sidebar_chip(cod, rol, selected, is_special=is_special,
                                          alertas=alertas, criticos=criticos)
            self._chip_refs[cod] = new_chip
        self._sidebar_chips.controls = list(self._chip_refs.values())

    def _refresh_theme_colors(self):
        """Actualiza colores de todos los componentes visuales sin recalcular datos."""
        raw = self._raw_data or {}
        alm_config = self._get_alm_config(raw)

        # Chips del sidebar
        self._refresh_chip_colors()
        self._refresh_chip_visibility()

        # Pills VENTA/MKTD
        if self._alm_view_pills and self._alm_view_pills.content:
            self._alm_view_pills.content.content.color = self.c["info"] if self._warehouse_group == "mktd" else self.c["accent"]
            try:
                self._alm_view_pills.content.update()
            except Exception:
                pass

        # Filtro salud
        if self._filtro_salud_row and self._filtro_salud_row.content:
            self._build_filtro_salud()
            try:
                self._filtro_salud_row.update()
            except Exception:
                pass

        # Cards de almacenes (solo reconstruye visuales, no datos)
        if self._warehouse_cards and self._kpis_alm:
            for card in self._card_instances:
                card.update_theme(self.c)

        # Sección categoría/líneas (solo reconstruye visuales, usa datos cached)
        if self._cat_section and self._kpis_alm and hasattr(self, '_lineas') and self._lineas:
            linea_section = LineaSection(self._lineas, self._categorias, self.c,
                                         on_linea_click=self._show_linea_skus,
                                         filtro_salud=self._filtro_salud,
                                         lineas_sin_catalogo=self._lineas_sin_catalogo,
                                         categorias_sin_catalogo=self._categorias_sin_catalogo)
            self._cat_section.content = linea_section.build()
            self._cat_section.visible = len(self._categorias) > 0 or len(self._categorias_sin_catalogo) > 0

        # KPI row
        if self._kpi_row and self._kpis_alm:
            total_disp = sum(a["disponible_total"] for a in self._kpis_alm.values())
            total_pre = sum(a["predespacho_total"] for a in self._kpis_alm.values())
            total_alertas = sum(a["alertas"] for a in self._kpis_alm.values())
            total_criticos = sum(a["criticos"] for a in self._kpis_alm.values())
            total_alto_stock = sum(a.get("alto_stock", 0) for a in self._kpis_alm.values())
            total_skus = sum(a["sku_count"] for a in self._kpis_alm.values())
            total_sin_catalogo = sum(a.get("sin_catalogo_count", 0) for a in self._kpis_alm.values())
            kpis = {
                "almacenes": len(self._kpis_alm), "skus": total_skus,
                "disponible": total_disp, "predespacho": total_pre,
                "alertas": total_alertas, "criticos": total_criticos,
                "alto_stock": total_alto_stock, "sin_catalogo": total_sin_catalogo,
            }
            self._kpi_row.content = self._build_kpi_row(kpis).content
            self._kpis_data = kpis

        # Transfer section (rebuild with new colors if visible)
        if self._transfer_section.visible:
            self._refresh_transfer_section()

        try:
            self.page.update()
        except Exception:
            pass

    def _refresh_transfer_section(self):
        """Reconstruye sección de transferencias con colores del tema actual."""
        if not self._raw_data:
            return
        search = (self._search_field.value or "").strip().lower()
        raw = self._raw_data if search else self._get_filtered_raw()
        alm_config = self._get_alm_config(raw)
        transfers = sugerir_transferencias(raw, alm_config, umbral=5, search=search)
        self._build_transfer_section(transfers, bool(transfers), bool(search))

    def _refresh_chip_colors(self):
        """Actualiza colores de chips existentes sin reconstruir datos."""
        raw = self._raw_data or {}
        alm_config = self._get_alm_config(raw)
        for cod, gd in self._chip_refs.items():
            cfg = alm_config.get(cod, {})
            rol = cfg.get("rol", "")
            selected = cod in self._selected_alms
            is_special = bool(SPECIAL_WAREHOUSE_RE.match(cod))
            rol_color = {"PRINCIPAL": self.c["accent"], "SECUNDARIO": self.c["info"], "EXTERNO": self.c["text_muted"]}
            base = rol_color.get(rol, "#666")
            inner = gd.content
            sel_border = ft.BorderSide(2, base) if selected else ft.BorderSide(0, "transparent")
            inner.bgcolor = rgba(base, 0.2) if selected else rgba(base, 0.05)
            inner.border = ft.Border(left=sel_border)
            row = inner.content
            for j, ctrl in enumerate(row.controls):
                if j == 0:
                    ctrl.bgcolor = base if selected else rgba(base, 0.6)
                elif j == 1:
                    ctrl.color = self.c["text_primary"] if selected else rgba(self.c["text_primary"], 0.5)
                elif j == len(row.controls) - 1:
                    ctrl.color = rgba(base, 0.6)
            # Update special icon
            if is_special and len(row.controls) > 2:
                for ctrl in row.controls[2:]:
                    if isinstance(ctrl, ft.Icon) and ctrl.tooltip and "MKTD" in ctrl.tooltip:
                        ctrl.color = rgba(self.c["text_muted"], 0.5)
            if hasattr(gd, 'content') and isinstance(gd.content, ft.Container):
                try:
                    gd.content.update()
                except Exception:
                    pass

    def _refresh_chip_visibility(self):
        """Ocultar/mostrar chips según el grupo activo."""
        raw = self._raw_data or {}
        alm_config = self._get_alm_config(raw)
        for cod, gd in self._chip_refs.items():
            is_mktd_chip = self._is_mktd(cod, alm_config)
            if (self._warehouse_group == "venta" and is_mktd_chip) or \
               (self._warehouse_group == "mktd" and not is_mktd_chip):
                gd.visible = False
            else:
                gd.visible = True

    def _update_sin_linea_footer(self, sin_linea: int):
        self._sin_cat_count = sin_linea
        if self._sin_cat_badge_text:
            self._sin_cat_badge_text.value = str(sin_linea) if sin_linea else ""
            self._sin_cat_badge.visible = bool(sin_linea)
        if self._sin_cat_row is not None:
            self._sin_cat_row.visible = bool(sin_linea)

    def _show_no_results(self, search: str):
        self._warehouse_cards.visible = False
        self._body_divider.visible = False
        self._cat_section.visible = False
        self._transfer_section.visible = False
        self._empty_state.content = self._build_empty_state(search)
        self._empty_state.visible = True
        self.page.update()

    def _update_empty_state(self, visible: bool):
        self._warehouse_cards.visible = not visible
        self._body_divider.visible = not visible
        self._cat_section.visible = not visible
        self._empty_state.visible = visible
        if visible:
            self._transfer_section.visible = False

    _TRANSFER_SORT = {
        "SKU": lambda t: t.get("sku", ""),
        "Producto": lambda t: (t.get("descripcion", "") or "").lower(),
        "Línea": lambda t: (t.get("linea", "") or "").lower(),
        "Alm": lambda t: t.get("almacen", ""),
        "Rol": lambda t: t.get("rol", ""),
        "Stock": lambda t: t.get("stock", 0),
        "Disp.": lambda t: t.get("disponible", 0),
        "VES": lambda t: t.get("p_disp", 0) / max(1, t.get("p_stock", 0)),
        "Origen": lambda t: t.get("secundario", ""),
        "Sugerencia": lambda t: (t.get("secundario", "") or "") + str(t.get("tipo", "")),
    }
    _TRANSFER_SORT_SUG = {
        "SKU": lambda t: t.get("sku", ""),
        "Producto": lambda t: (t.get("descripcion", "") or "").lower(),
        "VES": lambda t: t.get("p_disp", 0) / max(1, t.get("p_stock", 0)),
        "Origen": lambda t: t.get("secundario", ""),
        "Disp.": lambda t: t.get("s_disp", 0),
        "Sugerencia": lambda t: (t.get("secundario", "") or "") + str(t.get("tipo", "")),
    }

    def _on_transfer_sort(self, label: str):
        if label == self._transfer_sort_key:
            if self._transfer_sort_rev:
                self._transfer_sort_key = None
                self._transfer_sort_rev = False
            else:
                self._transfer_sort_rev = True
        else:
            self._transfer_sort_key = label
            self._transfer_sort_rev = False
        self._apply_filters()

    def _build_search_transfer_rows(self, transfers: list) -> tuple[list, str, str]:
        t_rows = [self._dlg_header([("SKU", 65), ("Producto", True), ("Línea", 65), ("Alm", 40, ft.TextAlign.CENTER),
                                    ("Rol", 55, ft.TextAlign.CENTER), ("Stock", 55, ft.TextAlign.RIGHT),
                                    ("Disp.", 60, ft.TextAlign.RIGHT)],
                                   sortable={"SKU", "Producto", "Línea", "Alm", "Rol", "Stock", "Disp."},
                                   on_sort=self._on_transfer_sort,
                                   active_label=self._transfer_sort_key,
                                   sort_reverse=self._transfer_sort_rev)]
        for t in transfers[:20]:
            linea_txt = t.get("linea", "") or "-"
            rol_color = {"PRINCIPAL": self.c["accent"], "SECUNDARIO": self.c["info"], "EXTERNO": self.c["text_muted"]}
            rc = rol_color.get(t["rol"], "#666")
            sc = self._sku_state(t["sku"], t['disponible'])
            t_rows.append(self._dlg_row([
                ft.Text(t["sku"], size=11, weight=ft.FontWeight.W_600, width=65),
                self._desc_text(t.get("descripcion", ""), 12, self.c["text_muted"]),
                ft.Text(linea_txt, size=11, width=65, color=self.c["text_muted"],
                        max_lines=1, overflow=ft.TextOverflow.ELLIPSIS, tooltip=linea_txt),
                ft.Text(t["almacen"], size=11, width=40, color=rc, text_align=ft.TextAlign.CENTER),
                ft.Text(t["rol"], size=10, width=55, color=rgba(rc, 0.7), text_align=ft.TextAlign.CENTER),
                ft.Text(f"{t['stock']:,}", size=11, width=55, text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{t['disponible']:,}", size=12, weight=ft.FontWeight.W_700, width=60, color=sc["color"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
            ]))
        count_txt = f"{len(transfers)} resultados"
        title = f"SKU: '{self._search_field.value}'"
        if self._transfer_sort_key:
            arrow = "▲" if not self._transfer_sort_rev else "▼"
            count_txt += f" · {arrow} {self._transfer_sort_key}"
        return t_rows, count_txt, title

    def _build_suggestion_transfer_rows(self, transfers: list) -> tuple[list, str, str]:
        t_rows = [self._dlg_header([("SKU", 65), ("Producto", True), ("VES", 65, ft.TextAlign.CENTER),
                                    ("Origen", 45, ft.TextAlign.CENTER), ("Disp.", 55, ft.TextAlign.RIGHT),
                                    ("Sugerencia", True)],
                                   sortable={"SKU", "Producto", "VES", "Origen", "Disp.", "Sugerencia"},
                                   on_sort=self._on_transfer_sort,
                                   active_label=self._transfer_sort_key,
                                   sort_reverse=self._transfer_sort_rev)]
        for t in transfers[:10]:
            accion = f"Liberar QC en {t['secundario']}" if t['secundario'] == "121" else f"Trasladar desde {t['secundario']}"
            tipo_icon = "⚠️" if t["tipo"] == "critico" else "⚖️"
            t_rows.append(self._dlg_row([
                ft.Text(t["sku"], size=11, weight=ft.FontWeight.W_600, width=65),
                self._desc_text(t.get("descripcion", ""), 12, self.c["text_muted"]),
                ft.Text(f"{t['p_disp']}/{t['p_pred']}", size=11, width=65, color=self.c["error"] if t['p_disp'] <= 2 else self.c["warning"], text_align=ft.TextAlign.CENTER, font_family=NUM_FONT,
                        tooltip=f"En {t['principal']} · Stock {t['p_stock']:,} · Pred. {t['p_pred']:,} · Disp. {t['p_disp']:,}"),
                ft.Text(t["secundario"], size=11, width=45, color=self.c["info"], text_align=ft.TextAlign.CENTER),
                ft.Text(f"{t['s_disp']:,}", size=12, weight=ft.FontWeight.W_700, width=55, color=self.c["success"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{tipo_icon} {accion}", size=12, color=self.c["text_muted"], expand=True),
            ]))
        count_txt = f"{len(transfers)} sugerencias"
        title = "Transferencias Sugeridas"
        if self._transfer_sort_key:
            arrow = "▲" if not self._transfer_sort_rev else "▼"
            count_txt += f" · {arrow} {self._transfer_sort_key}"
        return t_rows, count_txt, title

    def _build_transfer_section(self, transfers: list, has_transfers: bool, search: bool):
        if self._transfer_sort_key:
            key_fn = (self._TRANSFER_SORT_SUG if not search else self._TRANSFER_SORT).get(self._transfer_sort_key)
            if key_fn:
                transfers = sorted(transfers, key=key_fn, reverse=self._transfer_sort_rev)
        # No transfers and not in search mode → hide completely
        if not has_transfers and not search:
            self._transfer_section.visible = False
            return

        # Auto-expandir cuando hay transferencias nuevas
        if not self._transfer_collapsed and has_transfers and not search:
            self._transfer_collapsed = False

        if search:
            t_rows, count_txt, title = self._build_search_transfer_rows(transfers)
        else:
            t_rows, count_txt, title = self._build_suggestion_transfer_rows(transfers)

        empty_note = ft.Container(height=0) if has_transfers else (
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.SEARCH_OFF, size=12, color=self.c["text_muted"]),
                    ft.Text(f"'{self._search_field.value}' no encontrado", size=10, color=self.c["text_muted"]),
                ], spacing=4),
                margin=ft.Margin(top=4, bottom=0, left=0, right=0),
            ) if search else ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.INFO_OUTLINE, size=12, color=self.c["text_muted"]),
                    ft.Text("No hay SKUs con stock bajo en almacén principal que tengan stock en secundarios", size=10, color=self.c["text_muted"]),
                ], spacing=4),
                margin=ft.Margin(top=4, bottom=0, left=0, right=0),
            )
        )
        chevron = ft.Icon(ft.Icons.EXPAND_MORE if self._transfer_collapsed else ft.Icons.EXPAND_LESS,
                          size=18, color=self.c["text_muted"],
                          animate_rotation=ft.Animation(200, ft.AnimationCurve.EASE_IN_OUT))

        def _toggle_transfer(e):
            self._transfer_collapsed = not self._transfer_collapsed
            self._apply_filters()

        body = ft.Container(
            content=ft.Column([ft.Column(t_rows, spacing=1), empty_note], spacing=0),
            visible=not self._transfer_collapsed,
        )
        self._transfer_section.content = ft.Container(
            content=ft.Column([
                ft.GestureDetector(
                    content=ft.Row([
                        ft.Icon(ft.Icons.SWAP_HORIZ, size=16, color=self.c["warning"]),
                        ft.Text(title, size=13, weight=ft.FontWeight.W_700, color=self.c["text_primary"]),
                        ft.Container(expand=True),
                        ft.Text(count_txt, size=10, color=self.c["text_muted"]),
                        chevron,
                    ], spacing=6),
                    on_tap=_toggle_transfer,
                    mouse_cursor=ft.MouseCursor.CLICK,
                ),
                ft.Container(height=6),
                body,
            ], spacing=0),
            padding=ft.Padding(left=20, right=20, top=10, bottom=10),
            bgcolor=self.c["surface"],
            border=ft.Border(bottom=ft.BorderSide(1, self.c["border"])),
        )
        self._transfer_section.visible = True

        sin_linea_txt = f" • {self._sin_cat_count} SKUs sin categoría" if self._sin_cat_count else ""
        search_txt = f" • filtro: '{self._search_field.value}'" if search else ""
        status_parts = f"Datos actualizados — {len(self._kpis_alm)} almacenes, {len(self._categorias)} categorías, {len(self._lineas)} líneas"
        self.status_text.value = f"{status_parts}{sin_linea_txt}{search_txt}"
        self.status_text.color = self.c["accent"]

    def _build_kpi_row(self, kpis: dict) -> ft.Container:
        if not kpis:
            kpis = {"almacenes": 0, "skus": 0, "sin_catalogo": 0, "alertas": 0, "criticos": 0, "alto_stock": 0, "disponible": 0, "predespacho": 0}
        self._kpis_data = kpis
        kc = self.c.get("kpis", {})
        col = self.c

        def card(label, value, icon, color, on_click, alert_when=False):
            return ft.Container(
                content=ft.GestureDetector(
                    content=ft.Row([
                        ft.Container(width=4, height=32, bgcolor=color, border_radius=2),
                        ft.Column([
                            ft.Text(f"{value:,}" if isinstance(value, int) else str(value),
                                    size=18, weight=ft.FontWeight.W_700, color=col["text_primary"]),
                            ft.Row([
                                ft.Icon(icon, size=10, color=color),
                                ft.Text(label, size=10, color=col["text_muted"], weight=ft.FontWeight.W_500),
                            ], spacing=4),
                        ], spacing=2, horizontal_alignment=ft.CrossAxisAlignment.START, expand=True),
                    ], spacing=12, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    on_tap=lambda e: self._handle_kpi_click(on_click),
                    mouse_cursor=ft.MouseCursor.CLICK,
                ),
                bgcolor=col["surface_variant"],
                border_radius=10,
                shadow=ft.BoxShadow(blur_radius=10, color=color, offset=ft.Offset(0, 0), spread_radius=2),
                padding=ft.Padding(left=10, right=14, top=10, bottom=10),
                expand=True,
            )

        row1 = [
            card("Almacenes", kpis.get("almacenes", 0), ft.Icons.WAREHOUSE, kc.get("almacenes", col["cyan"]), self._show_almacenes_dlg),
            card("SKUs", kpis.get("skus", 0), ft.Icons.INVENTORY_2, kc.get("skus", col["violet"]), self._show_skus_dlg),
            card("Disponible", kpis.get("disponible", 0), ft.Icons.CHECK_CIRCLE, kc.get("disponible", col["cyan"]), self._show_disp_dlg,
                 alert_when=kpis.get("disponible", 0) < 100),
            card("Predespacho", kpis.get("predespacho", 0), ft.Icons.CALL_MADE, kc.get("predespacho", col["info"]), self._show_pred_dlg,
                 alert_when=kpis.get("predespacho", 0) > 1000),
        ]
        row2 = [
            card("Sin Cat.", kpis.get("sin_catalogo", 0), ft.Icons.HELP_OUTLINE, kc.get("sin_catalogo", col["pink"]), self._show_sin_catalogo_dlg,
                 alert_when=kpis.get("sin_catalogo", 0) > 0),
            card("Alertas", kpis.get("alertas", 0), ft.Icons.WARNING_AMBER, kc.get("alertas", col["warning"]), self._show_alertas_dlg,
                 alert_when=kpis.get("alertas", 0) > 0),
            card("Críticos", kpis.get("criticos", 0), ft.Icons.ERROR_OUTLINE, kc.get("criticos", col["error"]), self._show_criticos_dlg,
                 alert_when=kpis.get("criticos", 0) > 0),
            card("Alto pred.", kpis.get("alto_stock", 0), ft.Icons.TRENDING_UP, kc.get("alto_stock", col["orange"]), self._show_alto_predespacho_dlg,
                 alert_when=kpis.get("alto_stock", 0) > 0),
        ]
        self._kpi_row_bg = ft.Container(
            content=ft.Column([
                ft.Row(row1, spacing=10, height=60, alignment=ft.MainAxisAlignment.CENTER),
                ft.Row(row2, spacing=10, height=60, alignment=ft.MainAxisAlignment.CENTER),
            ], spacing=10),
            padding=ft.Padding(left=20, right=20, top=12, bottom=12),
            bgcolor=col["surface"],
            border=ft.Border(bottom=ft.BorderSide(1, col["border"])),
        )
        return self._kpi_row_bg

    def _show_snack(self, msg: str, is_error: bool = False):
        self.page.snack_bar = ft.SnackBar(
            content=ft.Row([
                ft.Icon(
                    ft.Icons.ERROR_OUTLINE if is_error else ft.Icons.CHECK_CIRCLE,
                    color="white", size=20
                ),
                ft.Text(msg, color="white", weight=ft.FontWeight.W_500, size=14),
            ], spacing=10),
            bgcolor=self.c["error"] if is_error else self.c["accent"],
            duration=4000,
            behavior=ft.SnackBarBehavior.FLOATING,
            width=500,
        )
        self.page.snack_bar.open = True
        self.page.update()

    def _handle_kpi_click(self, on_click):
        if not on_click:
            return
        if not self._raw_data:
            self._show_snack("Presione 'Actualizar' para cargar datos")
            return
        try:
            on_click()
        except Exception as ex:
            self._show_snack(f"Error: {ex}")

    # ─── Interactive Dialogs ───

    def _show_almacenes_dlg(self):
        if not self._kpis_alm:
            return
        items = []
        for d in self._kpis_alm.values():
            items.append((d["codigo"], d["stock_total"], d["predespacho_total"],
                          d["disponible_total"], d["alertas"], d["criticos"]))

        def build(d, i=0):
            cod, stock, pred, disp, alertas, criticos = d
            return self._dlg_row([
                ft.Text(cod, weight=ft.FontWeight.W_700, size=12, width=50),
                ft.Text(f"{stock:,}", size=11, width=80, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{pred:,}", size=11, width=80, color=rgba(self.c["warning"], 0.8), text_align=ft.TextAlign.CENTER, font_family=NUM_FONT),
                ft.Text(f"{disp:,}", size=12, weight=ft.FontWeight.W_700, width=90, color=self.c["success"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(str(alertas), size=11, width=60, color=self.c["warning"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(str(criticos), size=11, width=60, color=self.c["error"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
            ], index=i)

        hdr = [self._dlg_header([("Cód", 50), ("Stock", 80, ft.TextAlign.RIGHT), ("Pred.", 80, ft.TextAlign.CENTER),
                                 ("Disp.", 90, ft.TextAlign.RIGHT), ("Alertas", 60, ft.TextAlign.RIGHT),
                                 ("Críticos", 60, ft.TextAlign.RIGHT)])]
        sort_cols = [("Cód", lambda x: x[0]), ("Stock", lambda x: x[1]), ("Pred.", lambda x: x[2]),
                     ("Disp.", lambda x: x[3]), ("Alertas", lambda x: x[4]), ("Críticos", lambda x: x[5])]
        self._show_paginated_dlg("Almacenes", hdr, items, build, sort_columns=sort_cols, height=350)

    def _show_skus_dlg(self):
        if not self._categorias:
            return
        items = [(c["categoria"], c["skus"], c["disponible"]) for c in self._categorias]

        def build(d, i=0):
            cat, skus, disp = d
            return self._dlg_row([
                ft.Text(cat, weight=ft.FontWeight.W_700, size=12, expand=True),
                ft.Text(f"{skus:,} SKUs", size=12, width=100, color=self.c["text_primary"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{disp:,}", size=11, width=100, color=self.c["success"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
            ], index=i)

        hdr = [self._dlg_header([("Categoría", True), ("SKUs", 100, ft.TextAlign.RIGHT), ("Disp.", 100, ft.TextAlign.RIGHT)])]
        sort_cols = [("Categoría", lambda x: x[0]), ("SKUs", lambda x: x[1]), ("Disp.", lambda x: x[2])]
        self._show_paginated_dlg("SKUs por Categoría", hdr, items, build, sort_columns=sort_cols, height=300)

    def _show_disp_dlg(self):
        if not self._lineas:
            return
        items = [(ln["nombre"], ln["disponible"], ln["stock"], ln["skus"]) for ln in self._lineas]

        def build(d, i=0):
            nombre, disp, stock, skus = d
            return self._dlg_row([
                ft.Text(nombre, size=12, expand=True),
                ft.Text(f"{disp:,}", size=12, weight=ft.FontWeight.W_700, width=100, color=self.c["success"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{stock:,}", size=11, width=90, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(str(skus), size=11, width=60, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
            ], index=i)

        hdr = [self._dlg_header([("Línea", True), ("Disponible", 100, ft.TextAlign.RIGHT),
                                 ("Stock", 90, ft.TextAlign.RIGHT), ("SKUs", 60, ft.TextAlign.RIGHT)])]
        sort_cols = [("Línea", lambda x: x[0]), ("Disponible", lambda x: x[1]),
                     ("Stock", lambda x: x[2]), ("SKUs", lambda x: x[3])]
        self._show_paginated_dlg("Disponible por Línea", hdr, items, build,
                                 sort_key="Disponible", sort_reverse=True, sort_columns=sort_cols, height=400)

    def _show_pred_dlg(self):
        if not self._lineas:
            return
        items = []
        for ln in self._lineas:
            total = ln["predespacho"] + ln["disponible"]
            ratio = (ln["predespacho"] / total * 100) if total > 0 else 0
            items.append((ln["nombre"], ln["predespacho"], ratio))

        def build(d, i=0):
            nombre, pred, ratio = d
            return self._dlg_row([
                ft.Text(nombre, size=12, expand=True),
                ft.Text(f"{pred:,}", size=12, weight=ft.FontWeight.W_700, width=100, color=self.c["warning"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{ratio:.0f}%", size=11, width=70, color=rgba(self.c["warning"], 0.7), text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
            ], index=i)

        hdr = [self._dlg_header([("Línea", True), ("Predespacho", 100, ft.TextAlign.RIGHT),
                                 ("Ratio", 70, ft.TextAlign.RIGHT)])]
        sort_cols = [("Línea", lambda x: x[0]), ("Predespacho", lambda x: x[1]), ("Ratio", lambda x: x[2])]
        self._show_paginated_dlg("Predespacho por Línea", hdr, items, build,
                                 sort_key="Predespacho", sort_reverse=True, sort_columns=sort_cols, height=400)

    def _show_alertas_dlg(self):
        raw = self._get_filtered_raw()
        if not raw or not self._kpis_alm:
            return
        items = []
        for alm, skus in raw.items():
            for sku, info in skus.items():
                disp = info.get("disponible", info["stock"] - info["predespacho"])
                st = self._sku_state(sku, disp)
                if st["nivel"] == "alerta":
                    items.append((sku, info.get("descripcion", ""), alm, info.get("stock", 0), info.get("predespacho", 0), disp, st["stock_minimo"]))
        if not items:
            return

        def build(d, i=0):
            sku, desc, alm, stock, pred, disp, smin = d
            st = self._sku_state(sku, disp)
            return self._dlg_row([
                ft.Text(sku, size=11, weight=ft.FontWeight.W_600, width=65),
                self._desc_text(desc, 11, self.c["text_muted"]),
                ft.Text(alm, size=11, width=55),
                ft.Text(f"{stock:,}", size=11, width=70, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{pred:,}", size=12, width=55, text_align=ft.TextAlign.CENTER, font_family=NUM_FONT),
                ft.Text(f"{disp:,}", size=13, weight=ft.FontWeight.W_700, width=60, color=self.c["warning"],
                        text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT,
                        tooltip=f"Mínimo {smin:,} · Stock {stock:,} · Pred. {pred:,}"),
            ], index=i)

        hdr = [self._dlg_header([("SKU", 65), ("Producto", True), ("Almacén", 55),
                                 ("Stock", 70, ft.TextAlign.RIGHT), ("Pred.", 55, ft.TextAlign.CENTER),
                                 ("Disp.", 60, ft.TextAlign.RIGHT)])]
        sort_cols = [("SKU", lambda x: x[0]), ("Producto", lambda x: (x[1] or "").lower()),
                     ("Almacén", lambda x: x[2]), ("Stock", lambda x: x[3]),
                     ("Pred.", lambda x: x[4]), ("Disp.", lambda x: x[5])]
        self._show_paginated_dlg("Productos en Alerta (≥ mínimo, < 2× mínimo)", hdr, items, build, sort_columns=sort_cols)

    def _show_sin_catalogo_dlg(self):
        raw = self._get_filtered_raw()
        if not raw or not self._kpis_alm:
            return
        items = []
        for alm, skus in raw.items():
            for sku, info in skus.items():
                if not _sku_info(sku).get("sin_catalogo"):
                    continue
                disp = info.get("disponible", max(0, info.get("stock", 0) - info.get("predespacho", 0)))
                items.append((sku, info.get("descripcion", ""), alm, info.get("stock", 0), info.get("predespacho", 0), disp))
        if not items:
            return

        def build(d, i=0):
            return self._dlg_row([
                ft.Text(d[0], size=11, weight=ft.FontWeight.W_600, width=65),
                self._desc_text(d[1], 11, self.c["text_muted"]),
                ft.Text(d[2], size=11, width=55),
                ft.Text(f"{d[3]:,}", size=11, width=70, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{d[4]:,}", size=12, width=55, color=rgba(self.c["warning"], 0.8), text_align=ft.TextAlign.CENTER, font_family=NUM_FONT),
                ft.Text(f"{d[5]:,}", size=13, weight=ft.FontWeight.W_700, width=60, color=self.c["warning"],
                        text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
            ], index=i)

        hdr = [self._dlg_header([("SKU", 65), ("Producto", True), ("Almacén", 55),
                                 ("Stock", 70, ft.TextAlign.RIGHT), ("Pred.", 55, ft.TextAlign.CENTER),
                                 ("Disp.", 60, ft.TextAlign.RIGHT)])]
        sort_cols = [("SKU", lambda x: x[0]), ("Producto", lambda x: (x[1] or "").lower()),
                     ("Almacén", lambda x: x[2]), ("Stock", lambda x: x[3]),
                     ("Pred.", lambda x: x[4]), ("Disp.", lambda x: x[5])]
        self._show_paginated_dlg("SKUs sin Catálogo", hdr, items, build, sort_columns=sort_cols)

    _PAGE_SIZE = 100

    def _next_page(self, title_base, header_rows, data_items, row_builder, page, height,
                   sort_key=None, sort_reverse=False, sort_columns=None):
        # Optimización: En lugar de cerrar, podríamos actualizar el contenido del diálogo.
        # Por ahora, mantenemos la consistencia del framework pero aseguramos limpieza.
        self.page.close(self._active_dlg)
        self._show_paginated_dlg(title_base, header_rows, data_items, row_builder, page, height,
                                  sort_key=sort_key, sort_reverse=sort_reverse, sort_columns=sort_columns)

    def _show_paginated_dlg(self, title_base: str, header_rows: list,
                            data_items: list, row_builder,
                            page: int = 0, height: int = 550,
                            sort_key=None, sort_reverse=False, sort_columns=None):
        try:
            return self._show_paginated_dlg_impl(title_base, header_rows, data_items, row_builder,
                                                  page, height, sort_key, sort_reverse, sort_columns)
        except Exception as ex:
            print(f"[Dialog Error] {ex}")
            import traceback
            traceback.print_exc()

    def _sort_data_items(self, data_items, sort_key, sort_reverse, sort_columns):
        if sort_columns and sort_key is not None:
            key_fn = None
            for label, fn in sort_columns:
                if label == sort_key:
                    key_fn = fn
                    break
            if key_fn:
                if any(isinstance(d, (list, tuple)) and d and d[0] == "sep" for d in data_items):
                    data_list = []
                    sep_item = None
                    group = []
                    for d in data_items:
                        if d[0] == "sep":
                            if sep_item is not None:
                                group.sort(key=key_fn, reverse=sort_reverse)
                                data_list.append(sep_item)
                                data_list.extend(group)
                            sep_item = d
                            group = []
                        else:
                            group.append(d)
                    if sep_item is not None:
                        group.sort(key=key_fn, reverse=sort_reverse)
                        data_list.append(sep_item)
                        data_list.extend(group)
                else:
                    data_list = sorted(data_items, key=key_fn, reverse=sort_reverse)
            else:
                data_list = list(data_items)
        else:
            data_list = list(data_items)
        return data_list

    def _build_pagination_controls(self, page, total_pages, sort_key, sort_reverse, sort_columns,
                                    title_base, header_rows, data_items, row_builder, height):
        pagination = []
        nav_spacing = 4

        if page > 0:
            pagination.append(ft.GestureDetector(
                content=ft.Container(
                    content=ft.Row([ft.Text("←", size=12, weight=ft.FontWeight.W_600),
                                    ft.Text("Anterior", size=12, weight=ft.FontWeight.W_500)], spacing=2),
                    padding=ft.Padding(left=10, right=10, top=6, bottom=6),
                    bgcolor=rgba(self.c["text_muted"], 0.06), border_radius=8,
                ),
                on_tap=lambda e, p=page - 1: self._next_page(
                    title_base, header_rows, data_items, row_builder, p, height,
                    sort_key=sort_key, sort_reverse=sort_reverse, sort_columns=sort_columns
                ),
                mouse_cursor=ft.MouseCursor.CLICK,
            ))

        max_visible = 7
        half = max_visible // 2
        p_start = max(0, min(page - half, total_pages - max_visible))
        p_end = min(total_pages, p_start + max_visible)
        for p in range(p_start, p_end):
            is_current = p == page
            pagination.append(ft.GestureDetector(
                content=ft.Container(
                    content=ft.Text(str(p + 1), size=13, weight=ft.FontWeight.W_700 if is_current else ft.FontWeight.W_500,
                                    color="white" if is_current else self.c["text_muted"]),
                    padding=ft.Padding(left=10, right=10, top=6, bottom=6),
                    bgcolor=self.c["accent"] if is_current else rgba(self.c["text_muted"], 0.05),
                    border_radius=8,
                    shadow=ft.BoxShadow(blur_radius=4, color=rgba(self.c["accent"], 0.3)) if is_current else None,
                ),
                on_tap=lambda e, p=p: self._next_page(
                    title_base, header_rows, data_items, row_builder, p, height,
                    sort_key=sort_key, sort_reverse=sort_reverse, sort_columns=sort_columns
                ),
                mouse_cursor=ft.MouseCursor.CLICK,
            ))

        if page + 1 < total_pages:
            pagination.append(ft.GestureDetector(
                content=ft.Container(
                    content=ft.Row([ft.Text("Siguientes", size=12, weight=ft.FontWeight.W_500),
                                    ft.Text("→", size=12, weight=ft.FontWeight.W_600)], spacing=2),
                    padding=ft.Padding(left=10, right=10, top=6, bottom=6),
                    bgcolor=rgba(self.c["text_muted"], 0.06), border_radius=8,
                ),
                on_tap=lambda e, p=page + 1: self._next_page(
                    title_base, header_rows, data_items, row_builder, p, height,
                    sort_key=sort_key, sort_reverse=sort_reverse, sort_columns=sort_columns
                ),
                mouse_cursor=ft.MouseCursor.CLICK,
            ))

        pagination_row = ft.Container(
            content=ft.Row([
                *pagination,
                ft.Container(height=8),
                ft.Text("Pág.", size=11, color=self.c["text_muted"]),
                ft.Container(
                    content=ft.Text(f"{page + 1} / {total_pages}", size=12,
                                    weight=ft.FontWeight.W_700, color=self.c["accent"],
                                    font_family=NUM_FONT),
                    padding=ft.Padding(left=6, right=6, top=4, bottom=4),
                    bgcolor=rgba(self.c["accent"], 0.08),
                    border_radius=6,
                ),
            ], spacing=nav_spacing, alignment=ft.MainAxisAlignment.CENTER),
            padding=ft.Padding(left=8, right=8, top=8, bottom=6),
        )
        return pagination_row

    def _show_paginated_dlg_impl(self, title_base: str, header_rows: list,
                                  data_items: list, row_builder,
                                  page: int = 0, height: int = 550,
                                  sort_key=None, sort_reverse=False, sort_columns=None):
        data_list = self._sort_data_items(data_items, sort_key, sort_reverse, sort_columns)

        def _header_h(hdr):
            try:
                ctrls = hdr.content.controls
            except Exception:
                return 30
            return 56 if (ctrls and isinstance(ctrls[0], ft.Column)) else 30

        headers_h = sum(_header_h(h) for h in header_rows)
        page_size = max(25, min(self._PAGE_SIZE, (height - (30 + headers_h + 36) - 14) // 26))

        total = len(data_list)
        start = page * page_size
        end = min(start + page_size, total)
        data_slice = data_list[start:end]

        total_pages = (total + page_size - 1) // page_size

        sort_labels = {lbl for lbl, _ in (sort_columns or [])}

        def _sort_tap(label):
            if label == sort_key:
                if sort_reverse:
                    self._next_page(title_base, header_rows, data_items, row_builder, 0, height,
                                    sort_key=None, sort_reverse=False, sort_columns=sort_columns)
                else:
                    self._next_page(title_base, header_rows, data_items, row_builder, 0, height,
                                    sort_key=label, sort_reverse=True, sort_columns=sort_columns)
            else:
                self._next_page(title_base, header_rows, data_items, row_builder, 0, height,
                                sort_key=label, sort_reverse=False, sort_columns=sort_columns)

        final_headers = []
        if sort_labels:
            for header_row in header_rows:
                if (hasattr(header_row, 'content') and hasattr(header_row.content, 'controls')
                        and header_row.content.controls
                        and all(isinstance(c, ft.Text) for c in header_row.content.controls)):
                    new_controls = []
                    for ctrl in header_row.content.controls:
                        label = ctrl.value
                        if label in sort_labels:
                            is_active = label == sort_key
                            new_controls.append(self._header_sort_ctrl(
                                label, is_active, _sort_tap,
                                ctrl.width if ctrl.width is not None else None,
                                bool(ctrl.expand), ctrl.text_align))
                        else:
                            txt = ft.Text(label, size=11, weight=ft.FontWeight.W_700,
                                          color=ctrl.color, no_wrap=True,
                                          text_align=ctrl.text_align)
                            if ctrl.expand:
                                txt.expand = True
                            elif ctrl.width is not None:
                                txt.width = ctrl.width
                            new_controls.append(txt)
                    final_headers.append(ft.Container(
                        content=ft.Row(new_controls, spacing=6),
                        bgcolor=header_row.bgcolor, border_radius=header_row.border_radius,
                        padding=ft.Padding(left=8, right=8, top=6, bottom=6),
                    ))
                else:
                    final_headers.append(header_row)
        else:
            final_headers = list(header_rows)

        export_btn = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.FILE_DOWNLOAD, size=16, color=self.c["accent"]),
                ft.Text("Exportar Excel", size=12, color=self.c["accent"], weight=ft.FontWeight.W_600),
            ], spacing=4, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            padding=ft.Padding(left=10, right=10, top=5, bottom=5),
            bgcolor=rgba(self.c["accent"], 0.08),
            border_radius=20,
            on_click=lambda _: self._show_export_config_dlg(title_base, data_list),
        )

        visible = [export_btn] + final_headers + [row_builder(d, i) for i, d in enumerate(data_slice)]
        pagination_row = self._build_pagination_controls(
            page, total_pages, sort_key, sort_reverse, sort_columns,
            title_base, header_rows, data_items, row_builder, height
        )
        visible.append(pagination_row)

        showing = f"Pág. {page + 1} de {total_pages}" if total_pages > 1 else str(total)
        if sort_key and sort_columns:
            arrow = "\u25b2" if not sort_reverse else "\u25bc"
            showing = f"{showing} · {arrow} {sort_key}"

        n_sep = sum(1 for d in data_slice if isinstance(d, (list, tuple)) and d and d[0] == "sep")
        est = (30 + headers_h + (len(data_slice) - n_sep) * 26 + n_sep * 38 + 36)
        content_h = est if est > height else None
        self._show_dlg(f"{title_base} — {showing}", visible, content_h)

    def _show_criticos_dlg(self):
        raw = self._get_filtered_raw()
        if not raw or not self._kpis_alm:
            return
        items = []
        for alm, skus in raw.items():
            for sku, info in skus.items():
                disp = info.get("disponible", info["stock"] - info["predespacho"])
                st = self._sku_state(sku, disp)
                if st["nivel"] == "critico":
                    items.append((sku, info.get("descripcion", ""), alm, info.get("stock", 0), info.get("predespacho", 0), disp, st["stock_minimo"]))
        if not items:
            return

        def build(d, i=0):
            sku, desc, alm, stock, pred, disp, smin = d
            st = self._sku_state(sku, disp)
            return self._dlg_row([
                ft.Text(sku, size=11, weight=ft.FontWeight.W_600, width=65),
                self._desc_text(desc, 11, self.c["text_muted"]),
                ft.Text(alm, size=11, width=55),
                ft.Text(f"{stock:,}", size=11, width=70, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{pred:,}", size=12, width=55, color=rgba(self.c["warning"], 0.8), text_align=ft.TextAlign.CENTER, font_family=NUM_FONT),
                ft.Text(f"{disp:,}", size=13, weight=ft.FontWeight.W_700, width=60, color=self.c["error"],
                        text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT,
                        tooltip=f"Mínimo {smin:,} · Stock {stock:,} · Pred. {pred:,}"),
            ], index=i)

        hdr = [self._dlg_header([("SKU", 65), ("Producto", True), ("Almacén", 55),
                                 ("Stock", 70, ft.TextAlign.RIGHT), ("Pred.", 55, ft.TextAlign.CENTER),
                                 ("Disp.", 60, ft.TextAlign.RIGHT)])]
        sort_cols = [("SKU", lambda x: x[0]), ("Producto", lambda x: (x[1] or "").lower()),
                     ("Almacén", lambda x: x[2]), ("Stock", lambda x: x[3]),
                     ("Pred.", lambda x: x[4]), ("Disp.", lambda x: x[5])]
        self._show_paginated_dlg("Productos Críticos (< mínimo configurado)", hdr, items, build, sort_columns=sort_cols)

    def _show_alto_predespacho_dlg(self):
        raw = self._get_filtered_raw()
        if not raw:
            return
        items = []
        for alm, skus in raw.items():
            for sku, info in skus.items():
                stock = info.get("stock", 0)
                pred = info.get("predespacho", 0)
                if stock > 0 and pred / stock >= 0.85:
                    ratio = round(pred / stock * 100, 1)
                    disp = info.get("disponible", max(0, stock - pred))
                    items.append((sku, info.get("descripcion", ""), alm, stock, pred, ratio, disp))
        if not items:
            return

        def build(d, i=0):
            sku, desc, alm, stock, pred, ratio, disp = d
            st = self._sku_state(sku, disp)
            return self._dlg_row([
                ft.Text(sku, size=11, weight=ft.FontWeight.W_600, width=65),
                self._desc_text(desc, 11, self.c["text_muted"]),
                ft.Text(alm, size=11, width=55),
                ft.Text(f"{stock:,}", size=12, width=70, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{pred:,}", size=12, width=55, text_align=ft.TextAlign.CENTER, font_family=NUM_FONT),
                ft.Text(f"{ratio}%", size=12, width=55, color=rgba(self.c["orange"], 0.9 if ratio >= 99 else 0.7),
                        text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT,
                        tooltip=f"Stock {stock:,} · Pred. {pred:,} · Disp. {disp:,} · {ratio}% predespacho"),
            ], index=i)

        hdr = [self._dlg_header([("SKU", 65), ("Producto", True), ("Almacén", 55),
                                 ("Stock", 70, ft.TextAlign.RIGHT), ("Pred.", 55, ft.TextAlign.CENTER),
                                 ("%", 55, ft.TextAlign.RIGHT)])]
        sort_cols = [("SKU", lambda x: x[0]), ("Producto", lambda x: (x[1] or "").lower()),
                     ("Almacén", lambda x: x[2]), ("Stock", lambda x: x[3]),
                     ("Pred.", lambda x: x[4]), ("%", lambda x: x[5])]
        self._show_paginated_dlg("SKUs alto predespacho", hdr, items, build, sort_columns=sort_cols)

    def _show_warehouse_skus(self, alm_data: dict):
        cod = alm_data["codigo"]
        stock = self._raw_data.get(cod, {}) if self._raw_data else {}
        raw_skus = sorted(stock.items(), key=lambda x: (
            _sku_info(x[0])["indice"],
            x[1].get("disponible", 0)
        ))
        if not raw_skus:
            return
        raw_skus = [(sku, info) for sku, info in raw_skus if self._match_filtro_sku(sku, info.get("disponible", info["stock"] - info["predespacho"]))]
        if not raw_skus:
            return

        grouped: dict[str, list] = {}
        for sku, info in raw_skus:
            cat = _sku_info(sku)["categoria"]
            grouped.setdefault(cat, []).append((sku, info))

        items = []
        for cat in sorted(grouped, key=lambda c: self._CAT_ORDER.get(c, 9)):
            cat_list = grouped[cat]
            cat_color = self._CAT_COLORS.get(cat, self.c["text_muted"])
            cat_icons = {"VINIBALL": ft.Icons.SPORTS_SOCCER, "VINIFAN": ft.Icons.COLORIZE,
                         "REPRESENTADAS": ft.Icons.BUSINESS, "OTROS": ft.Icons.HELP_OUTLINE}
            items.append(("sep", f"{cat} ({len(cat_list)})", cat_color, cat_icons.get(cat, "")))
            for sku, info in cat_list:
                items.append(("row", sku, info, _sku_info(sku)["indice"]))

        total_pred = sum(info["predespacho"] for _, info in raw_skus)
        total_disp = sum(info.get("disponible", info["stock"] - info["predespacho"]) for _, info in raw_skus)
        summary = self._summary_bar([
            ("SKUs", f"{len(raw_skus)}", self.c["text_primary"]),
            ("Predespacho", f"{total_pred:,}", self.c["warning"]),
            ("Disponible", f"{total_disp:,}", self.c["success"]),
        ])

        data_idx = [0]

        def row_builder(d, _i=0):
            nonlocal data_idx
            if d[0] == "sep":
                return self._dlg_separator(d[1], d[2], d[3])
            sku, info, cat_idx = d[1], d[2], d[3]
            disp = info.get("disponible", info["stock"] - info["predespacho"])
            st = self._sku_state(sku, disp)
            un_bx = _sku_info(sku)["un_bx"]
            bx = info["stock"] // un_bx if un_bx > 0 else info["stock"]
            dbx = disp // un_bx if un_bx > 0 else disp
            idx = data_idx[0]
            data_idx[0] += 1
            dot_color = st["color"]
            dot = ft.Container(width=10, height=10, border_radius=5, bgcolor=dot_color,
                               shadow=ft.BoxShadow(blur_radius=6, color=rgba(dot_color, 0.5)))
            return self._dlg_row([
                ft.Text(str(cat_idx), size=10, color=self.c["text_muted"], width=35, text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(sku, size=11, weight=ft.FontWeight.W_600, width=65),
                self._desc_text(info.get("descripcion", ""), 12, self.c["text_muted"]),
                ft.Text(info.get("sku_unit", ""), size=11, width=30, color=self.c["text_muted"], text_align=ft.TextAlign.CENTER),
                ft.Text(f"{bx:,}", size=12, width=50, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{info['predespacho']:,}", size=12, width=55, color=rgba(self.c["warning"], 0.8), text_align=ft.TextAlign.CENTER, font_family=NUM_FONT),
                ft.Text(f"{disp:,}", size=13, weight=ft.FontWeight.W_700, width=60, color=st["color"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{dbx:,}", size=12, width=50, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                dot,
            ], index=idx)

        hdr = [summary, self._dlg_header([("#", 35, ft.TextAlign.RIGHT), ("SKU", 65), ("Producto", True),
                                          ("Und", 30, ft.TextAlign.CENTER), ("BX", 50, ft.TextAlign.RIGHT),
                                          ("Pred.", 55, ft.TextAlign.CENTER), ("Disp.", 60, ft.TextAlign.RIGHT),
                                          ("D.BX", 50, ft.TextAlign.RIGHT), ("", 10)])]
        sort_cols = [
            ("#", lambda x: x[3]),
            ("SKU", lambda x: x[1]),
            ("Producto", lambda x: (x[2].get("descripcion", "") or "").lower()),
            ("Und", lambda x: x[2].get("sku_unit", "") or ""),
            ("BX", lambda x: x[2].get("stock", 0) // _sku_info(x[1])["un_bx"] if _sku_info(x[1])["un_bx"] > 0 else x[2].get("stock", 0)),
            ("Pred.", lambda x: x[2].get("predespacho", 0)),
            ("Disp.", lambda x: x[2].get("disponible", x[2].get("stock", 0) - x[2].get("predespacho", 0))),
            ("D.BX", lambda x: x[2].get("disponible", x[2].get("stock", 0) - x[2].get("predespacho", 0)) // _sku_info(x[1])["un_bx"] if _sku_info(x[1])["un_bx"] > 0 else x[2].get("disponible", 0)),
        ]
        self._show_paginated_dlg(f"{cod} — Detalle SKUs", hdr, items, row_builder, sort_columns=sort_cols)

    def _show_linea_skus(self, linea_cod: str):
        raw = self._get_filtered_raw()
        if not raw:
            return
        seen = set()
        items = []
        for alm, all_skus in raw.items():
            for sku, info in all_skus.items():
                if sku in seen:
                    continue
                seen.add(sku)
                cat_info = _sku_info(sku)
                if cat_info["linea"] == linea_cod:
                    disp = info.get("disponible", info["stock"] - info["predespacho"])
                    if not self._match_filtro_sku(sku, disp):
                        continue
                    items.append((sku, info.get("descripcion", ""), info.get("sku_unit", ""),
                                 info["stock"], info["predespacho"], disp, alm,
                                 cat_info.get("indice", 9999)))

        if not items:
            return
        items.sort(key=lambda x: x[7])  # Orden inicial por índice

        nombre_linea = linea_cod
        for ln in (self._lineas or []):
            if ln["codigo"] == linea_cod:
                nombre_linea = ln.get("nombre", linea_cod)
                break

        def build(d, i=0):
            sku, desc, unit, st, pre, disp, alm, idx = d
            un_bx = _sku_info(sku)["un_bx"]
            bx = st // un_bx if un_bx > 0 else st
            dbx = disp // un_bx if un_bx > 0 else disp
            sc = self._sku_state(sku, disp)
            dot_color = sc["color"]
            dot = ft.Container(width=10, height=10, border_radius=5, bgcolor=dot_color,
                               shadow=ft.BoxShadow(blur_radius=6, color=rgba(dot_color, 0.5)))
            return self._dlg_row([
                ft.Text(str(idx), size=10, color=self.c["text_muted"], width=35, text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(sku, size=11, weight=ft.FontWeight.W_600, width=65),
                self._desc_text(desc, 11, self.c["text_muted"]),
                ft.Text(unit, size=11, width=30, color=self.c["text_muted"], text_align=ft.TextAlign.CENTER),
                ft.Text(f"{bx:,}", size=12, width=50, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{pre:,}", size=12, width=55, color=rgba(self.c["warning"], 0.8), text_align=ft.TextAlign.CENTER, font_family=NUM_FONT),
                ft.Text(f"{disp:,}", size=13, weight=ft.FontWeight.W_700, width=60, color=sc["color"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(f"{dbx:,}", size=12, width=50, color=self.c["text_muted"], text_align=ft.TextAlign.RIGHT, font_family=NUM_FONT),
                ft.Text(alm, size=11, width=40),
                dot,
            ], index=i)

        hdr = [self._dlg_header([("#", 35, ft.TextAlign.RIGHT), ("SKU", 65), ("Producto", True),
                                 ("Und", 30, ft.TextAlign.CENTER), ("BX", 50, ft.TextAlign.RIGHT),
                                 ("Pred.", 55, ft.TextAlign.CENTER), ("Disp.", 60, ft.TextAlign.RIGHT),
                                 ("D.BX", 50, ft.TextAlign.RIGHT), ("Alm", 40), ("", 10)])]
        sort_cols = [
            ("#", lambda x: x[7]),
            ("SKU", lambda x: x[0]),
            ("Producto", lambda x: (x[1] or "").lower()),
            ("Und", lambda x: x[2] or ""),
            ("BX", lambda x: x[3] // _sku_info(x[0])["un_bx"] if _sku_info(x[0])["un_bx"] > 0 else x[3]),
            ("Pred.", lambda x: x[4]),
            ("Disp.", lambda x: x[5]),
            ("D.BX", lambda x: x[5] // _sku_info(x[0])["un_bx"] if _sku_info(x[0])["un_bx"] > 0 else x[5]),
            ("Alm", lambda x: x[6]),
        ]
        self._show_paginated_dlg(f"{nombre_linea} — SKUs", hdr, items, build, sort_columns=sort_cols)



    def _show_sku_editor_dlg(self, sku: str, desc: str):
        config = load_lineas()
        lineas_list = config.get("lineas", [])
        linea_options = [ft.dropdown.Option(ln["codigo"], ln.get("nombre", ln["codigo"])) for ln in lineas_list]
        for ln in self._lineas:
            if ln["codigo"] not in {opt.key for opt in linea_options}:
                linea_options.append(ft.dropdown.Option(ln["codigo"], ln.get("nombre", ln["codigo"])))

        linea_dd = ft.Dropdown(
            label="Línea", options=linea_options,
            dense=True, width=350, text_size=13,
        )
        cat_options = [ft.dropdown.Option(c) for c in sorted(self._categorias, key=lambda x: x["categoria"])]
        if "OTROS" not in [c["categoria"] for c in self._categorias]:
            cat_options.append(ft.dropdown.Option("OTROS"))
        cat_dd = ft.Dropdown(
            label="Categoría",
            options=cat_options,
            dense=True, width=350, text_size=13,
        )

        def on_linea_change(e):
            cat = ""
            for meta in get_api_sku_meta().values():
                if meta.get("linea") == linea_dd.value and meta.get("categoria"):
                    cat = meta["categoria"]
                    break
            if not cat:
                cat = self._LINEA_CAT.get(linea_dd.value, "")
            if cat:
                cat_dd.value = cat
                self.page.update()

        linea_dd.on_change = on_linea_change

        async def on_save(e):
            linea = linea_dd.value
            categoria = cat_dd.value
            if not linea:
                return
            update_catalogo_sku(sku, desc, linea, categoria)
            self.page.close(dlg)
            self._show_snack(f"SKU {sku} asignado a {linea}")
            self._apply_filters()

        dlg = ft.AlertDialog(
            title=ft.Row([ft.Icon(ft.Icons.EDIT_SQUARE, color=self.c["accent"]), ft.Text(f"Editar SKU {sku}", weight=ft.FontWeight.W_800, size=16)]),
            content=ft.Container(content=ft.Column([
                ft.Text(f"Descripción: {desc or '—'}", size=12, color=self.c["text_muted"]),
                ft.Container(height=8),
                linea_dd, cat_dd,
            ], spacing=6, tight=True), width=400),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda _: self.page.close(dlg)),
                ft.ElevatedButton("Guardar", bgcolor=self.c["accent_dark"], color="white", on_click=on_save,
                                  style=ft.ButtonStyle(overlay_color=rgba(self.c["accent"], 0.1), elevation=0)),
            ],
        )
        self.page.open(dlg)

    # ─── Dialog Helpers ───

    def _sku_state(self, sku: str, disp: int) -> dict:
        cat_info = _sku_info(sku)
        un_bx = cat_info["un_bx"]
        stock_minimo = self._lineas_config.get(cat_info["linea"], {}).get("stock_minimo", 0) if cat_info["linea"] else 0
        umbral = stock_minimo if stock_minimo > 0 else un_bx
        if umbral > 0 and disp < umbral:
            return {"nivel": "critico", "color": self.c["error"], "emoji": "🔴", "stock_minimo": stock_minimo}
        if umbral > 0 and disp < umbral * 2:
            return {"nivel": "alerta", "color": self.c["warning"], "emoji": "🟡", "stock_minimo": stock_minimo}
        return {"nivel": "ok", "color": self.c["success"], "emoji": "🟢", "stock_minimo": stock_minimo}

    def _match_filtro_sku(self, sku: str, disp: int) -> bool:
        cat_info = _sku_info(sku)
        un_bx = cat_info["un_bx"]
        stock_minimo = self._lineas_config.get(cat_info["linea"], {}).get("stock_minimo", 0) if cat_info["linea"] else 0
        umbral = stock_minimo if stock_minimo > 0 else un_bx
        if umbral > 0 and disp < umbral:
            estado = "critico"
        elif umbral > 0 and disp < umbral * 2:
            estado = "alerta"
        else:
            estado = "ok"
        if self._filtro_salud == "todo":
            return True
        return estado == self._filtro_salud

    _CAT_ORDER = {"VINIBALL": 0, "VINIFAN": 1, "REPRESENTADAS": 2, "INDUSTRIAL": 3, "INDUMENTARIA": 4,
                  "PUBLICIDAD": 5, "MATERIALES": 6, "DESCARTE Y VARIOS": 7, "CIPTECH": 8, "PRODUCCION": 9,
                  "OTROS": 10}
    _CAT_COLORS = {
        "VINIBALL": ACCENT,
        "VINIFAN": "#3b82f6",
        "REPRESENTADAS": "#8b5cf6",
        "INDUSTRIAL": "#f59e0b",
        "INDUMENTARIA": "#ec4899",
        "PUBLICIDAD": "#06b6d4",
        "MATERIALES": "#84cc16",
        "DESCARTE Y VARIOS": "#78716c",
        "CIPTECH": "#14b8a6",
        "PRODUCCION": "#a855f7",
        "OTROS": "#6b7280",
    }
    _LINEA_CAT = {
        "PELOTAS": "VINIBALL", "MASCOTAS": "VINIBALL",
        "ACCESORIOS": "VINIFAN", "ARCHIVO": "VINIFAN", "DIBUJO": "VINIFAN",
        "DIDACTICOS": "VINIFAN", "ESCRITURA": "VINIFAN", "FORROS": "VINIFAN",
        "MANUALIDADES": "VINIFAN", "METALICA": "VINIFAN", "PEGAMENTOS": "VINIFAN",
        "PINTURA": "VINIFAN", "REPRESENTADAS": "REPRESENTADAS",
    }

    def _header_txt(self, label: str, is_active: bool, width=None,
                    text_align=None) -> ft.Text:
        txt = ft.Text(label, size=11,
                      weight=ft.FontWeight.W_800 if is_active else ft.FontWeight.W_700,
                      color=self.c["accent"], no_wrap=True)
        if width is not None:
            txt.width = width
        if text_align is not None:
            txt.text_align = text_align
        txt.tooltip = f"Ordenar por {label}"
        return txt

    def _header_sort_ctrl(self, label: str, is_active: bool,
                          on_sort, width=None, expand=False, text_align=None):
        txt = self._header_txt(label, is_active, width, text_align)
        gd = ft.GestureDetector(
            content=txt,
            on_tap=lambda e: on_sort(label),
            mouse_cursor=ft.MouseCursor.CLICK,
        )
        if expand:
            gd.expand = True
        return gd

    def _dlg_header(self, columns: list[tuple], sortable: set = None,
                    on_sort=None, active_label: str = None,
                    sort_reverse: bool = False) -> ft.Container:
        sortable = sortable or set()
        controls = []
        for col in columns:
            label, width = col[0], col[1]
            align = col[2] if len(col) > 2 else ft.TextAlign.LEFT
            is_active = label in sortable and label == active_label
            if label in sortable and on_sort:
                controls.append(self._header_sort_ctrl(
                    label, is_active, on_sort,
                    width if isinstance(width, int) else None,
                    width is True, align))
            else:
                txt = ft.Text(label, size=11, weight=ft.FontWeight.W_700,
                              color=self.c["accent"], no_wrap=True, text_align=align)
                if width is True:
                    txt.expand = True
                elif width is not None:
                    txt.width = width
                controls.append(txt)
        return ft.Container(
            content=ft.Row(controls, spacing=6),
            bgcolor=rgba(self.c["accent"], 0.12), border_radius=6,
            padding=ft.Padding(left=8, right=8, top=6, bottom=6),
        )

    def _dlg_row(self, controls: list, index: int = 0) -> ft.Container:
        bg = rgba(self.c["accent"], 0.03) if index % 2 == 1 else "transparent"
        return ft.Container(
            content=ft.Row(controls, spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor=bg, padding=ft.Padding(left=8, right=8, top=3, bottom=3),
        )

    def _desc_text(self, value, size: int = 11, color=None, weight=None, expand=True) -> ft.Text:
        text = (value or "").strip()
        return ft.Text(text, size=size, color=color, weight=weight, expand=expand,
                       max_lines=1, overflow=ft.TextOverflow.ELLIPSIS, tooltip=text or None)

    def _dlg_separator(self, text: str, color: str, icon: str = "") -> ft.Container:
        return ft.Container(
            content=ft.Row([
                ft.Icon(icon, size=14, color=color) if icon else ft.Container(width=0),
                ft.Text(text, size=13, weight=ft.FontWeight.W_700, color=color),
            ], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor=rgba(color, 0.08),
            border_radius=6,
            padding=ft.Padding(left=10, right=10, top=6, bottom=6),
        )

    def _summary_bar(self, items: list[tuple[str, str, str]]) -> ft.Container:
        cells = []
        for label, value, color in items:
            cells.append(ft.Column([
                ft.Text(label, size=10, color=self.c["text_muted"], weight=ft.FontWeight.W_500),
                ft.Text(value, size=14, weight=ft.FontWeight.W_800, color=color),
            ], spacing=1, horizontal_alignment=ft.CrossAxisAlignment.CENTER))
        return ft.Container(
            content=ft.Row(cells, spacing=24, alignment=ft.MainAxisAlignment.CENTER),
            bgcolor=rgba(self.c["accent"], 0.06), border_radius=8,
            padding=ft.Padding(left=12, right=12, top=8, bottom=8),
        )

    def _build_empty_state(self, search: str) -> ft.Container:
        return ft.Container(
            content=ft.Column([
                ft.Icon(ft.Icons.SEARCH_OFF, size=42, color=rgba(self.c["text_muted"], 0.4)),
                ft.Text(f"No se encontraron resultados para '{search}'",
                        size=14, color=self.c["text_muted"], weight=ft.FontWeight.W_600,
                        text_align=ft.TextAlign.CENTER),
                ft.Text("Revisá el código SKU o probá con otra búsqueda",
                        size=12, color=rgba(self.c["text_muted"], 0.7), text_align=ft.TextAlign.CENTER),
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
            alignment=ft.alignment.center,
            padding=ft.Padding(left=24, right=24, top=56, bottom=56),
        )

    def _show_dlg(self, title: str, rows: list, height: int | None):
        dlg = ft.AlertDialog(
            title=ft.Text(title, weight=ft.FontWeight.W_800, size=16, color=self.c["text_primary"]),
            content=ft.Container(
                ft.Column(rows, spacing=0, scroll=ft.ScrollMode.AUTO if height is not None else None),
                width=760,
                height=height,
            ),
            actions=[ft.TextButton("Cerrar", on_click=lambda e: self.page.close(dlg))],
        )
        self._active_dlg = dlg
        self.page.open(dlg)


    def update_data(self, raw_data: dict[str, dict[str, dict]], cache_timestamp: str | None = None, api_timestamp: str | None = None, stale: bool = False):
        self._raw_data = raw_data
        self._cache_timestamp = cache_timestamp
        self._stale_data = stale
        self._stale_badge.visible = stale
        self._update_refresh_status(cache_timestamp, api_timestamp, stale)
        alm_config = self._get_alm_config(raw_data)

        # Default selection based on API category data
        venta_codes = {c for c in raw_data.keys() if self._is_venta(c, alm_config, raw_data)}
        mktd_codes = {c for c in raw_data.keys() if self._is_mktd(c, alm_config, raw_data)}
        self._selected_alms = venta_codes.copy()
        self._warehouse_group = "venta"

        sort_order = {"PRINCIPAL": 0, "SECUNDARIO": 1, "EXTERNO": 2}
        sorted_codes = sorted(raw_data.keys(), key=lambda c: (sort_order.get(alm_config.get(c, {}).get("rol", ""), 9), alm_config.get(c, {}).get("prioridad", 99)))

        # Separar en dos grupos
        venta_chips: list[ft.Container] = []
        mktd_chips: list[ft.Container] = []
        self._chip_refs: dict[str, ft.GestureDetector] = {}
        kpis = getattr(self, "_kpis_alm", None) or {}
        for cod in sorted_codes:
            rol = alm_config.get(cod, {}).get("rol", "")
            is_special = bool(SPECIAL_WAREHOUSE_RE.match(cod))
            selected = cod in self._selected_alms
            alertas = kpis.get(cod, {}).get("alertas", 0)
            criticos = kpis.get(cod, {}).get("criticos", 0)
            is_mktd = self._is_mktd(cod, alm_config)
            chip = self._sidebar_chip(cod, rol, selected, is_special=is_special,
                                      alertas=alertas, criticos=criticos, is_mktd=is_mktd)
            self._chip_refs[cod] = chip
            if is_mktd:
                mktd_chips.append(chip)
            else:
                venta_chips.append(chip)

        # Construir columna de chips con separadores de sección
        chips_section = []
        if self._warehouse_group != "mktd":
            chips_section.append(ft.Container(
                content=ft.Text("VENTA", size=9, weight=ft.FontWeight.W_700, color=self.c["accent"]),
                padding=ft.Padding(left=12, right=12, top=8, bottom=2),
            ))
            chips_section.extend(venta_chips)
        if mktd_chips and self._warehouse_group != "venta":
            if chips_section:
                chips_section.append(ft.Divider(height=1, color=rgba(self.c["border"], 0.5)))
            chips_section.append(ft.Container(
                content=ft.Text("MKTD", size=9, weight=ft.FontWeight.W_700, color=self.c["info"]),
                padding=ft.Padding(left=12, right=12, top=4, bottom=2),
            ))
            chips_section.extend(mktd_chips)

        self._sidebar_chips.controls = chips_section
        self._apply_filters()

    def set_offline(self, offline: bool):
        self._offline_badge.visible = offline
        if offline:
            self._offline_badge.opacity = 0
            self._offline_badge.update()
            self._offline_badge.opacity = 1
        self.page.update()

    def _set_refreshing(self, refreshing: bool):
        self._refreshing = refreshing
        if not self._refresh_status_badge:
            return
        row = self._refresh_status_badge.content
        dot, txt = row.controls
        if refreshing:
            dot.icon = ft.Icons.CIRCLE
            dot.color = self.c["accent"]
            txt.value = "Actualizando..."
            txt.color = self.c["accent"]
            self._refresh_status_badge.visible = True
            self._refresh_status_badge.bgcolor = rgba(self.c["accent"], 0.08)
        elif self._cache_timestamp:
            self._update_refresh_status(self._cache_timestamp, self._api_timestamp, self._stale_data)
        self._refresh_status_badge.update()

    def _format_ts_display(self, ts: str | None) -> str:
        if not ts:
            return "-"
        try:
            dt = datetime.fromisoformat(ts)
            return dt.strftime("%H:%M")
        except Exception:
            return str(ts)[:16]

    def _get_api_cache_ttl(self) -> tuple[int, bool]:
        """Retorna (ttl_min, cache_expirado) desde el API meta."""
        try:
            from src.core.s1_downloader import get_api_meta
            meta = get_api_meta()
            ttl_sec = meta.get("cache_expiro_en") or 900
            expired = meta.get("cache_expirado", False)
            return int(ttl_sec / 60), bool(expired)
        except Exception:
            from src.core.constants import AUTO_REFRESH_INTERVAL
            return int(AUTO_REFRESH_INTERVAL / 60), False

    def _update_refresh_status(self, cache_timestamp: str | None, api_timestamp: str | None = None, stale: bool = False):
        if not self._refresh_status_badge:
            return
        self._stale_data = stale
        self._cache_timestamp = cache_timestamp
        self._api_timestamp = api_timestamp or self._api_timestamp
        row = self._refresh_status_badge.content
        dot, txt = row.controls
        if self._refreshing:
            return
        if not cache_timestamp:
            self._refresh_status_badge.visible = False
            return
        try:
            ts = datetime.fromisoformat(cache_timestamp)
            if ts.tzinfo is not None:
                ts = ts.replace(tzinfo=None)
            cache_age_min = (datetime.now().replace(tzinfo=None) - ts).total_seconds() / 60
        except Exception:
            self._refresh_status_badge.visible = False
            return

        from src.core.constants import AUTO_REFRESH_INTERVAL

        # Get real API cache metadata
        ttl_min, cache_expired = self._get_api_cache_ttl()

        age_min = cache_age_min
        if api_timestamp:
            try:
                api_ts = datetime.fromisoformat(api_timestamp)
                if api_ts.tzinfo is not None:
                    api_ts = api_ts.replace(tzinfo=None)
                api_age = (datetime.now().replace(tzinfo=None) - api_ts).total_seconds() / 60
                age_min = min(api_age, cache_age_min)
            except Exception:
                pass

        remaining = max(0, ttl_min - int(age_min))
        is_fresh = age_min <= ttl_min * 0.5
        is_warn = age_min >= ttl_min and not stale  # expired but auto-refresh not yet triggered
        is_stale = stale or cache_expired or age_min > ttl_min * 2

        if is_stale:
            dot.color = self.c["error"]
            txt.value = f"{remaining} min"
            txt.color = self.c["error"]
            self._refresh_status_badge.bgcolor = rgba(self.c["error"], 0.08)
            self._refresh_status_badge.tooltip = (
                f"CACHE EXPIRADO\n"
                f"API: {self._format_ts_display(api_timestamp)}\n"
                f"App: {self._format_ts_display(cache_timestamp)}\n"
                f"Tiempo restante: {remaining} min\n"
                f"Auto-refresh: cada {ttl_min} min\n"
                f"Click para forzar actualización."
            )
        elif is_warn:
            dot.color = self.c["warning"]
            txt.value = f"{remaining} min"
            txt.color = self.c["warning"]
            self._refresh_status_badge.bgcolor = rgba(self.c["warning"], 0.08)
            self._refresh_status_badge.tooltip = (
                f"CACHE VENCIDO — esperando nuevo\n"
                f"API: {self._format_ts_display(api_timestamp)}\n"
                f"App: {self._format_ts_display(cache_timestamp)}\n"
                f"Tiempo restante: {remaining} min\n"
                f"Auto-refresh: cada {ttl_min} min\n"
                f"Click para forzar actualización."
            )
        else:
            dot.color = self.c["success"]
            txt.value = f"{remaining} min" if not is_fresh else "OK"
            txt.color = self.c["success"]
            self._refresh_status_badge.bgcolor = rgba(self.c["success"], 0.08)
            status_label = "Fresco" if is_fresh else "Cache"
            self._refresh_status_badge.tooltip = (
                f"CACHE {status_label.upper()}\n"
                f"API: {self._format_ts_display(api_timestamp)}\n"
                f"App: {self._format_ts_display(cache_timestamp)}\n"
                f"Tiempo restante: {remaining} min\n"
                f"Auto-refresh: cada {ttl_min} min\n"
                f"Click para forzar actualización."
            )
        self._refresh_status_badge.visible = True
        self._refresh_status_badge.update()

    def _show_stale_warning(self):
        self._stale_badge.visible = True
        self._ts_badge.bgcolor = rgba(self.c["warning"], 0.07)
        self._stale_badge.update()

    def _hide_stale_warning(self):
        self._stale_badge.visible = False
        self._ts_badge.bgcolor = rgba(self.c["accent"], 0.07)
        self._stale_badge.update()

    def _update_health_badge(self):
        if not self._health_badge or not self._kpis_alm:
            return
        total_criticos = sum(a.get("criticos", 0) for a in self._kpis_alm.values())
        total_alertas = sum(a.get("alertas", 0) for a in self._kpis_alm.values())
        text = self._health_badge.content
        if total_criticos > 0:
            text.controls[0].value = f"{total_criticos} críticos"
            text.controls[0].color = self.c["error"]
            self._health_badge.bgcolor = rgba(self.c["error"], 0.08)
            self._health_badge.border = ft.Border(
                top=ft.BorderSide(1, rgba(self.c["error"], 0.3)),
                right=ft.BorderSide(1, rgba(self.c["error"], 0.3)),
                bottom=ft.BorderSide(1, rgba(self.c["error"], 0.3)),
                left=ft.BorderSide(1, rgba(self.c["error"], 0.3)),
            )
            self._health_badge.visible = True
        elif total_alertas > 0:
            text.controls[0].value = f"{total_alertas} alertas"
            text.controls[0].color = self.c["warning"]
            self._health_badge.bgcolor = rgba(self.c["warning"], 0.08)
            self._health_badge.border = ft.Border(
                top=ft.BorderSide(1, rgba(self.c["warning"], 0.3)),
                right=ft.BorderSide(1, rgba(self.c["warning"], 0.3)),
                bottom=ft.BorderSide(1, rgba(self.c["warning"], 0.3)),
                left=ft.BorderSide(1, rgba(self.c["warning"], 0.3)),
            )
            self._health_badge.visible = True
        else:
            self._health_badge.visible = False

    def format_cache_timestamp(self, timestamp_iso: str | None) -> str:
        if not timestamp_iso:
            return ""
        try:
            ts = datetime.fromisoformat(timestamp_iso)
            now = datetime.now()
            delta = now - ts
            if delta.total_seconds() < 60:
                return " (cache 0m)"
            if delta.total_seconds() < 3600:
                mins = int(delta.total_seconds() / 60)
                return f" (cache {mins}m)"
            hours = int(delta.total_seconds() / 3600)
            return f" (cache {hours}h)"
        except Exception:
            return " (cache)"

    def _show_export_config_dlg(self, title: str, data: list):
        config = load_lineas()
        alm_config = self._get_alm_config(self._raw_data or {})
        has_skus = bool(_extract_report_skus(data, self._raw_data or {}))
        ext_codes = sorted(c for c, cfg in alm_config.items() if cfg.get("rol") == "EXTERNO")

        scope_radio = ft.RadioGroup(content=ft.Column([
            ft.Radio(value="control", label="Control (VES + Secundarios + QC)"),
            ft.Radio(value="todos", label="Todos (incluye externos)" + (f" {', '.join(ext_codes)}" if ext_codes else "")),
        ]), value="control")
        mode_radio = ft.RadioGroup(content=ft.Column([
            ft.Radio(value="basic", label="Básico (Solo Disponibilidad)"),
            ft.Radio(value="detailed", label="Completo (Stock + Predespacho)"),
        ]), value="basic")
        summary_check = ft.Checkbox(
            label="Incluir Resumen por línea (VES · QC · Sec. · Ext. · Salud)",
            value=True,
        )

        def on_confirm(_):
            detailed = mode_radio.value == "detailed"
            summary = summary_check.value
            scope = scope_radio.value

            def on_result(e: ft.FilePickerResultEvent):
                path = e.path
                if not path:
                    return
                try:
                    export_to_excel(
                        data, path, title,
                        detailed, summary,
                        scope=scope,
                        raw=self._raw_data,
                        alm_config=alm_config,
                        lineas_config=config.get("lineas", []),
                    )
                    self._show_snack("Archivo Excel generado con éxito")
                    import os
                    os.startfile(path)
                except Exception as ex:
                    self._show_snack(f"Error al exportar: {ex}", is_error=True)

            self.page.close(dlg)
            self._file_picker.on_result = on_result
            self._file_picker.save_file(file_name=f"{_make_report_name(title)}.xlsx")

        content_cols = []
        if has_skus:
            content_cols.append(ft.Text("Alcance del reporte", size=11, weight=ft.FontWeight.W_700, color=self.c["text_muted"]))
            content_cols.append(scope_radio)
        content_cols.append(ft.Text("Nivel de detalle", size=11, weight=ft.FontWeight.W_700, color=self.c["text_muted"]))
        content_cols.append(mode_radio)
        content_cols.append(summary_check)

        dlg = ft.AlertDialog(
            title=ft.Row([ft.Icon(ft.Icons.SETTINGS_SUGGEST, color=self.c["accent"]), ft.Text("Configurar Exportación", weight=ft.FontWeight.W_800, size=16)]),
            content=ft.Container(content=ft.Column([
                ft.Container(
                    content=ft.Column(content_cols, spacing=8),
                    padding=10,
                ),
            ], spacing=5, tight=True), width=420),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda _: self.page.close(dlg)),
                ft.ElevatedButton("Generar Archivo", bgcolor=self.c["accent_dark"], color="white", on_click=on_confirm,
                                  style=ft.ButtonStyle(overlay_color=rgba(self.c["accent"], 0.1), elevation=0)),
            ],
        )
        self.page.open(dlg)


