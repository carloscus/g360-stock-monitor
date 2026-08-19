from __future__ import annotations

import json
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from src.core.constants import LINEAS_FILE, DATA_DIR
from src.core.s1_downloader import get_api_sku_meta


APP_AUTHOR = "g360-stock-monitor"
APP_NAME = "G360"


def _make_report_name(title: str) -> str:
    """Genera nombre de archivo con timestamp: G360_{slug}_{YYYYMMDD}_{HHMMSS}.xlsx"""
    now = datetime.now()
    ts = now.strftime("%Y%m%d_%H%M%S")
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", title).strip("_").upper()
    if len(slug) > 40:
        slug = slug[:40]
    if not slug:
        slug = "REPORTE"
    return f"{APP_NAME}_{slug}_{ts}"


_CATLOGO: list[dict] | None = None
_SKU_INDEX: dict[str, dict] | None = None


def _build_sku_index() -> dict[str, dict]:
    """Construye un diccionario SKU → entry para búsquedas O(1)."""
    global _SKU_INDEX
    if _SKU_INDEX is not None:
        return _SKU_INDEX
    _SKU_INDEX = {}
    for p in _load_catalogo():
        sku = str(p.get("sku", "")).strip()
        if sku:
            _SKU_INDEX[sku] = p
    return _SKU_INDEX


def _load_catalogo() -> list[dict]:
    global _CATLOGO
    if _CATLOGO is not None:
        return _CATLOGO
    path = DATA_DIR / "catalogo_productos.json"
    if not path.exists():
        _CATLOGO = []
        return _CATLOGO
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    _CATLOGO = data.get("productos", [])
    return _CATLOGO

def reload_catalogo():
    """Limpia la caché del catálogo y el índice para forzar recarga."""
    global _CATLOGO, _SKU_INDEX
    _CATLOGO = None
    _SKU_INDEX = None
    _load_catalogo()
    _build_sku_index()

def update_catalogo_sku(sku: str, descripcion: str = "", linea: str = "", categoria: str = ""):
    """Agrega o actualiza un SKU en catalogo_productos.json y recarga la caché."""
    global _CATLOGO, _SKU_INDEX
    path = DATA_DIR / "catalogo_productos.json"
    data = {"productos": []}
    if path.exists():
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    productos = data.get("productos", [])
    found = None
    for p in productos:
        if str(p.get("sku", "")).strip() == str(sku).strip():
            found = p
            break
    if found:
        if linea:
            found["linea"] = linea
        if categoria:
            found["categoria"] = categoria
        if descripcion:
            found["descripcion"] = descripcion
    else:
        entry = {"sku": sku, "orden": 9999, "un_bx": 1}
        if linea:
            entry["linea"] = linea
        if categoria:
            entry["categoria"] = categoria
        if descripcion:
            entry["descripcion"] = descripcion
        productos.append(entry)
    data["productos"] = productos
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    _CATLOGO = None
    _SKU_INDEX = None
    _load_catalogo()
    _build_sku_index()

def _sku_info(sku: str) -> dict:
    """Usa metadata del API como fuente de verdad. Catálogo local solo como respaldo para info/reportes."""
    meta = get_api_sku_meta().get(sku)
    if meta is not None:
        return {
            "linea": meta.get("linea", ""),
            "linea_nombre": meta.get("linea_nombre", ""),
            "categoria": meta.get("categoria", ""),
            "un_bx": meta.get("un_bx", 1),
            "precio_lista": meta.get("precio_lista", 0),
            "sin_catalogo": meta.get("sin_catalogo", False),
            "estado_linea": meta.get("estado_linea", ""),
            "indice": meta.get("orden", 9999),
        }
    # Fallback solo para info/reportes: busca en catálogo local
    idx = _build_sku_index()
    cat_entry = idx.get(sku)
    if cat_entry:
        return {
            "linea": cat_entry.get("linea", ""),
            "linea_nombre": "",
            "categoria": cat_entry.get("categoria", ""),
            "un_bx": cat_entry.get("un_bx", 1),
            "precio_lista": cat_entry.get("precio_lista", 0),
            "sin_catalogo": False,
            "estado_linea": "",
            "indice": cat_entry.get("orden", 9999),
        }
    return {"linea": "", "linea_nombre": "", "categoria": "OTROS", "un_bx": 1, "precio_lista": 0, "sin_catalogo": False, "estado_linea": "", "indice": 9999}


def load_lineas() -> dict:
    with open(LINEAS_FILE, encoding="utf-8") as f:
        return json.load(f)


def _warehouse_disp(warehouse_data: dict[str, dict], sku: str) -> int:
    info = warehouse_data.get(sku)
    if not info:
        return 0
    return info.get("disponible", max(0, info.get("stock", 0) - info.get("predespacho", 0)))


def _process_sku_kpi(sku: str, info: dict, stock_minimo: int = 0, un_bx: int = 1,
                     linea_cod: str = "", categoria: str = "", sin_catalogo: bool = False) -> dict:
    stock = info.get("stock", 0)
    pred = info.get("predespacho", 0)
    disp = info.get("disponible", max(0, stock - pred))
    umbral = stock_minimo if stock_minimo > 0 else un_bx
    return {
        "stock": stock,
        "pred": pred,
        "disp": disp,
        "un_bx": un_bx,
        "linea_cod": linea_cod,
        "categoria": categoria,
        "sin_catalogo": sin_catalogo,
        "critico": umbral > 0 and disp < umbral,
        "alerta": umbral > 0 and umbral <= disp < umbral * 2,
        "alto_stock": stock > 0 and pred / stock >= 0.85,
        "stock_minimo": stock_minimo,
    }


def _update_linea_kpi(lineas_kpi: dict[str, dict], linea_cod: str, sku_kpi: dict):
    if not linea_cod:
        return
    if linea_cod not in lineas_kpi:
        lineas_kpi[linea_cod] = {"skus": 0, "stock": 0, "predespacho": 0, "disponible": 0, "alto_stock": 0}
    lineas_kpi[linea_cod]["skus"] += 1
    lineas_kpi[linea_cod]["stock"] += sku_kpi["stock"]
    lineas_kpi[linea_cod]["predespacho"] += sku_kpi["pred"]
    lineas_kpi[linea_cod]["disponible"] += sku_kpi["disp"]
    if sku_kpi["alto_stock"]:
        lineas_kpi[linea_cod]["alto_stock"] += 1


def _update_categoria_kpi(categorias_kpi: dict[str, dict], categoria: str, sku_kpi: dict):
    if not categoria:
        return
    if categoria not in categorias_kpi:
        categorias_kpi[categoria] = {"skus": 0, "stock": 0, "predespacho": 0, "disponible": 0}
    categorias_kpi[categoria]["skus"] += 1
    categorias_kpi[categoria]["stock"] += sku_kpi["stock"]
    categorias_kpi[categoria]["predespacho"] += sku_kpi["pred"]
    categorias_kpi[categoria]["disponible"] += sku_kpi["disp"]


def calcular_kpis_almacen(
    raw: dict[str, dict[str, dict]]
) -> dict[str, dict]:
    config = load_lineas()
    lineas_config = {ln["codigo"]: ln for ln in config.get("lineas", [])}
    result = {}
    for cod_alm, skus in raw.items():
        lineas_kpi: dict[str, dict] = {}
        categorias_kpi: dict[str, dict] = {}
        stock_total = 0
        predespacho_total = 0
        disponible_total = 0
        bx_total = 0
        disp_bx = 0
        alertas = 0
        criticos = 0
        alto_stock = 0
        sku_count = 0
        sin_catalogo_count = 0

        for sku, info in skus.items():
            cat_info = _sku_info(sku)
            stock_minimo = lineas_config.get(cat_info["linea"], {}).get("stock_minimo", 0)
            un_bx = cat_info["un_bx"]
            sku_kpi = _process_sku_kpi(
                sku, info,
                stock_minimo=stock_minimo,
                un_bx=un_bx,
                linea_cod=cat_info["linea"],
                categoria=cat_info["categoria"],
                sin_catalogo=cat_info.get("sin_catalogo", False),
            )
            stock = sku_kpi["stock"]
            pred = sku_kpi["pred"]
            disp = sku_kpi["disp"]

            stock_total += stock
            predespacho_total += pred
            disponible_total += disp
            sku_count += 1

            bx_total += stock // un_bx if un_bx > 0 else stock
            disp_bx += disp // un_bx if un_bx > 0 else disp

            if sku_kpi["critico"]:
                criticos += 1
            elif sku_kpi["alerta"]:
                alertas += 1

            if sku_kpi["alto_stock"]:
                alto_stock += 1

            if sku_kpi["sin_catalogo"]:
                sin_catalogo_count += 1

            _update_linea_kpi(lineas_kpi, sku_kpi["linea_cod"], sku_kpi)
            _update_categoria_kpi(categorias_kpi, sku_kpi["categoria"], sku_kpi)

        prev = _load_snapshot(cod_alm)
        cambio = None
        if prev and prev.get("disponible_total") is not None:
            diff = disponible_total - prev["disponible_total"]
            pct = (diff / prev["disponible_total"] * 100) if prev["disponible_total"] > 0 else 0
            cambio = {"absoluto": diff, "porcentaje": round(pct, 1)}

        result[cod_alm] = {
            "codigo": cod_alm,
            "stock_total": stock_total,
            "predespacho_total": predespacho_total,
            "disponible_total": disponible_total,
            "bx_total": bx_total,
            "disponible_bx": disp_bx,
            "sku_count": sku_count,
            "sin_catalogo_count": sin_catalogo_count,
            "alertas": alertas,
            "criticos": criticos,
            "alto_stock": alto_stock,
            "lineas": lineas_kpi,
            "categorias": categorias_kpi,
            "cambio": cambio,
            "ultima_actualizacion": datetime.now().strftime("%H:%M"),
        }
        _save_snapshot(cod_alm, {"disponible_total": disponible_total, "timestamp": datetime.now().isoformat()})

    return result


def obtener_metricas_lineas(
    kpis_por_almacen: dict[str, dict],
    raw_data: dict[str, dict[str, dict]] | None = None,
    alm_config: dict | None = None,
) -> tuple[list[dict], list[dict]]:
    config = load_lineas()
    lineas_config = {ln["codigo"]: ln for ln in config.get("lineas", [])}
    linea_categoria: dict[str, str] = {}
    linea_nombre: dict[str, str] = {}
    for meta in get_api_sku_meta().values():
        lc = meta.get("linea", "")
        cat = meta.get("categoria", "")
        if lc and cat:
            linea_categoria[lc] = cat
        if lc and meta.get("linea_nombre"):
            linea_nombre[lc] = meta.get("linea_nombre", "")
    for p in _load_catalogo():
        lc = p.get("linea", "")
        cat = p.get("categoria", "")
        if lc and cat and lc not in linea_categoria:
            linea_categoria[lc] = cat
        if lc and lc not in linea_nombre:
            linea_nombre[lc] = lc

    lineas_aggregated: dict[str, dict] = {}
    lineas_sin_catalogo: dict[str, dict] = {}

    for alm_data in kpis_por_almacen.values():
        cod_alm = alm_data.get("codigo", "")
        cfg_alm = alm_config.get(cod_alm, {}) if alm_config else {}
        rol = cfg_alm.get("rol", "")

        for cod_linea, info in alm_data.get("lineas", {}).items():
            target = lineas_aggregated
            if raw_data and cod_alm in raw_data:
                tiene_vigente = any(
                    not _sku_info(sku).get("sin_catalogo")
                    for sku in raw_data[cod_alm]
                    if _sku_info(sku).get("linea") == cod_linea
                )
                if not tiene_vigente:
                    target = lineas_sin_catalogo

            if cod_linea not in target:
                cfg = lineas_config.get(cod_linea, {})
                target[cod_linea] = {
                    "codigo": cod_linea,
                    "nombre": cfg.get("nombre", linea_nombre.get(cod_linea, cod_linea)),
                    "categoria": linea_categoria.get(cod_linea, ""),
                    "grupo": cfg.get("grupo", ""),
                    "stock_minimo": cfg.get("stock_minimo", 0),
                    "stock": 0,
                    "predespacho": 0,
                    "disponible": 0,
                    "skus": 0,
                    "alto_stock": 0,
                    "stock_ves": 0,
                    "stock_qc": 0,
                    "stock_secundario": 0,
                    "stock_externo": 0,
                    "predespacho_ves": 0,
                    "estado_linea": "",
                }
            entry = target[cod_linea]
            entry["stock"] += info["stock"]
            entry["predespacho"] += info["predespacho"]
            entry["disponible"] += info["disponible"]
            entry["skus"] += info["skus"]
            entry["alto_stock"] += info.get("alto_stock", 0)
            if not entry["estado_linea"]:
                entry["estado_linea"] = info.get("estado_linea", "")

            if rol == "PRINCIPAL":
                entry["stock_ves"] += info["stock"]
                entry["predespacho_ves"] += info["predespacho"]
            elif cod_alm == "121":
                entry["stock_qc"] += info["stock"]
            elif rol == "SECUNDARIO":
                entry["stock_secundario"] += info["stock"]
            elif rol == "EXTERNO":
                entry["stock_externo"] += info["stock"]

    for entry in list(lineas_aggregated.values()) + list(lineas_sin_catalogo.values()):
        sm = entry["stock_minimo"]
        sv = entry["stock_ves"]
        if sm > 0 and sv < sm:
            entry["salud"] = "critico"
            entry["pct_minimo"] = round(sv / sm * 100, 1)
        elif sm > 0 and sv < sm * 2:
            entry["salud"] = "alerta"
            entry["pct_minimo"] = round(sv / sm * 100, 1)
        else:
            entry["salud"] = "bueno"
            entry["pct_minimo"] = round(sv / sm * 100, 1) if sm > 0 else 100

    return (
        sorted(lineas_aggregated.values(), key=lambda x: x["disponible"], reverse=True),
        sorted(lineas_sin_catalogo.values(), key=lambda x: x["disponible"], reverse=True),
    )


def obtener_metricas_categorias(
    kpis_por_almacen: dict[str, dict],
    raw_data: dict[str, dict[str, dict]] | None = None,
) -> tuple[list[dict], list[dict]]:
    categorias: dict[str, dict] = {}
    categorias_sin_catalogo: dict[str, dict] = {}

    for alm_data in kpis_por_almacen.values():
        cod_alm = alm_data.get("codigo", "")
        for cat, info in alm_data.get("categorias", {}).items():
            if not cat:
                continue
            target = categorias
            if raw_data and cod_alm in raw_data:
                tiene_vigente = any(
                    not _sku_info(sku).get("sin_catalogo")
                    for sku in raw_data.get(cod_alm, [])
                    if _sku_info(sku).get("categoria") == cat
                )
                if not tiene_vigente:
                    target = categorias_sin_catalogo

            if cat not in target:
                target[cat] = {"categoria": cat, "stock": 0, "predespacho": 0, "disponible": 0, "skus": 0}
            target[cat]["stock"] += info["stock"]
            target[cat]["predespacho"] += info["predespacho"]
            target[cat]["disponible"] += info["disponible"]
            target[cat]["skus"] += info["skus"]

    orden = {"VINIBALL": 0, "VINIFAN": 1, "REPRESENTADAS": 2}
    return (
        sorted(categorias.values(), key=lambda x: (orden.get(x["categoria"], 99), x["categoria"])),
        sorted(categorias_sin_catalogo.values(), key=lambda x: (orden.get(x["categoria"], 99), x["categoria"])),
    )


def _snapshot_path(warehouse_code: str) -> str:
    return str(DATA_DIR / f"_snapshot_{warehouse_code}.json")


def _load_snapshot(warehouse_code: str) -> dict | None:
    path = _snapshot_path(warehouse_code)
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None
    return None


def _save_snapshot(warehouse_code: str, data: dict):
    path = _snapshot_path(warehouse_code)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def leer_ultima_actualizacion() -> str | None:
    """Timestamp más reciente entre los snapshots de almacenes, en formato HH:MM:SS.

    Devuelve None si no hay snapshots válidos.
    """
    latest = None
    for path in DATA_DIR.glob("_snapshot_*.json"):
        try:
            with open(path, encoding="utf-8") as f:
                ts = json.load(f).get("timestamp")
            if ts and (latest is None or ts > latest):
                latest = ts
        except Exception:
            continue
    if not latest:
        return None
    try:
        return datetime.fromisoformat(latest).strftime("%H:%M:%S")
    except Exception:
        return str(latest)[:19]


def _warehouse_sort_key(cod: str, alm_config: dict) -> tuple:
    cfg = alm_config.get(cod, {})
    rol = cfg.get("rol", "")
    rol_order = {"PRINCIPAL": 0, "SECUNDARIO": 1, "EXTERNO": 2}
    return (rol_order.get(rol, 9), cfg.get("prioridad", 99))


def _build_search_transfer_result(sku: str, info: dict, almacen: str, alm_config: dict) -> dict:
    d = info
    disp = d.get("disponible", max(0, d.get("stock", 0) - d.get("predespacho", 0)))
    cat = _sku_info(sku)
    rol = alm_config.get(almacen, {}).get("rol", "")
    return {
        "sku": sku,
        "descripcion": d.get("descripcion", ""),
        "linea": cat["linea"],
        "un_bx": cat["un_bx"],
        "almacen": almacen,
        "rol": rol,
        "stock": d.get("stock", 0),
        "predespacho": d.get("predespacho", 0),
        "disponible": disp,
    }


def _add_transfer_match(resultados: list[dict], sku: str, pinfo: dict, p_disp: int,
                         scod: str, sdata: dict, s_disp: int, tipo: str, principal: str):
    cat = _sku_info(sku)
    resultados.append({
        "sku": sku,
        "descripcion": pinfo.get("descripcion", ""),
        "linea": cat["linea"],
        "un_bx": cat["un_bx"],
        "tipo": tipo,
        "principal": principal,
        "p_stock": pinfo.get("stock", 0),
        "p_pred": pinfo.get("predespacho", 0),
        "p_disp": p_disp,
        "secundario": scod,
        "s_stock": sdata.get("stock", 0),
        "s_pred": sdata.get("predespacho", 0),
        "s_disp": s_disp,
    })


def sugerir_transferencias(
    raw: dict[str, dict[str, dict]],
    alm_config: dict,
    umbral: int = 5,
    search: str = ""
) -> list[dict]:
    principal = None
    secundarios = []
    for cod, cfg in alm_config.items():
        rol = cfg.get("rol", "")
        if rol == "PRINCIPAL":
            principal = cod
        elif rol == "SECUNDARIO":
            secundarios.append(cod)

    if search:
        resultados = []
        seen_skus = set()
        for cod_alm, skus in raw.items():
            for sku, info in skus.items():
                desc = (info.get("descripcion", "") or "").lower()
                if search not in sku.lower() and search not in desc:
                    continue
                if sku in seen_skus:
                    continue
                seen_skus.add(sku)
                for a in sorted(raw.keys(), key=lambda c: _warehouse_sort_key(c, alm_config)):
                    d = raw[a].get(sku)
                    if not d:
                        continue
                    resultados.append(_build_search_transfer_result(sku, d, a, alm_config))
        return resultados[:20]

    if not principal:
        return []
    p_data = raw.get(principal, {})
    resultados = []

    for sku, pinfo in p_data.items():
        p_stock = pinfo.get("stock", 0)
        p_pred = pinfo.get("predespacho", 0)
        p_disp = pinfo.get("disponible", max(0, p_stock - p_pred))

        for scod in sorted(secundarios, key=lambda c: _warehouse_sort_key(c, alm_config)):
            sdata = raw.get(scod, {}).get(sku)
            if not sdata:
                continue
            s_stock = sdata.get("stock", 0)
            s_pred = sdata.get("predespacho", 0)
            s_disp = sdata.get("disponible", max(0, s_stock - s_pred))

            if p_disp <= umbral and s_disp > 0:
                _add_transfer_match(resultados, sku, pinfo, p_disp, scod, sdata, s_disp, "critico", principal)
                break

            if s_stock > 0 and p_stock > 0 and (s_stock / p_stock) >= 3 and p_disp > umbral:
                _add_transfer_match(resultados, sku, pinfo, p_disp, scod, sdata, s_disp, "desbalance", principal)
                break

    return resultados

def _qc_codes(alm_config: dict) -> set:
    """Códigos de almacén de control de calidad (nombre contiene 'inspección')."""
    return {cod for cod, cfg in alm_config.items()
            if "INSPECCION" in (cfg.get("nombre") or "").upper()}


def _extract_report_skus(data_items: list, raw: dict | None) -> list:
    """Extrae los SKUs reales presentes en data_items, verificados contra raw."""
    skus = []
    for d in data_items:
        if not isinstance(d, (list, tuple)) or not d:
            continue
        if d[0] == "sep":
            continue
        sku = d[1] if d[0] == "row" and len(d) > 1 else d[0]
        if not isinstance(sku, str) or not sku.isdigit():
            continue
        if raw is not None and not any(isinstance(al, dict) and sku in al for al in raw.values()):
            continue
        skus.append(sku)
    return list(dict.fromkeys(skus))


def _row_fields(d) -> dict | None:
    """Normaliza un item de data_items a {sku, desc, unit, st, pre, disp} o None."""
    if not isinstance(d, (list, tuple)) or not d:
        return None
    if d[0] == "row" and len(d) >= 3:
        info = d[2] or {}
        cat = _sku_info(d[1])
        return {
            "sku": d[1],
            "desc": info.get("descripcion", "") or cat.get("descripcion", ""),
            "unit": cat.get("sku_unit") or "UND",
            "st": info.get("stock", 0),
            "pre": info.get("predespacho", 0),
            "disp": info.get("disponible", max(0, info.get("stock", 0) - info.get("predespacho", 0))),
        }
    sku = d[0]
    if not isinstance(sku, str) or not sku.isdigit():
        return None
    n = len(d)
    if n >= 8:  # (sku, desc, unit, st, pre, disp, alm, idx) — SKUs por Línea
        return {"sku": sku, "desc": d[1], "unit": d[2] or "UND",
                "st": d[3], "pre": d[4], "disp": d[5]}
    if n == 7:  # (sku, desc, alm, pred, disp, ratio, stock) — ≥85%
        return {"sku": sku, "desc": d[1], "unit": "UND", "st": d[6], "pre": d[3], "disp": d[4]}
    if n == 6:  # (sku, desc, unit/alm, st, pre, disp) — clásico o alertas/críticos
        cat = _sku_info(sku)
        return {"sku": sku, "desc": d[1], "unit": cat.get("sku_unit") or "UND",
                "st": d[3], "pre": d[4], "disp": d[5]}
    if n == 5:  # (sku, desc, alm, pred, disp) — sin categoría
        return {"sku": sku, "desc": d[1], "unit": "UND", "st": d[3] + d[4], "pre": d[3], "disp": d[4]}
    if n == 4:  # (sku, desc, alm, disp) — alertas / críticos
        return {"sku": sku, "desc": d[1], "unit": "UND", "st": 0, "pre": 0, "disp": d[3]}
    return {"sku": sku, "desc": str(d[1]) if n > 1 else "", "unit": "UND", "st": 0, "pre": 0, "disp": 0}


def _build_role_report(skus: list, raw: dict, alm_config: dict, lineas_config: list | None, scope: str) -> dict:
    """Agrega stock por línea y por rol (VES / QC / Secundario / Externo) dentro del alcance."""
    qc = _qc_codes(alm_config)
    lc = {l["codigo"]: l for l in (lineas_config or [])}
    wh = {cod: cfg for cod, cfg in alm_config.items()
          if cod in raw and (scope != "control" or cfg.get("participa_control"))}
    aggr = {}
    for cod, cfg in wh.items():
        rol = cfg.get("rol", "")
        is_qc = cod in qc
        for sku in skus:
            info = raw[cod].get(sku)
            if not info:
                continue
            cat = _sku_info(sku)
            linea = cat.get("linea") or ""
            e = aggr.get(linea)
            if e is None:
                cl = lc.get(linea, {})
                e = aggr[linea] = {
                    "nombre": cl.get("nombre") or cat.get("linea_nombre") or linea,
                    "stock_minimo": cl.get("stock_minimo", 0),
                    "skus": set(), "stock": 0, "pred": 0,
                    "ves": 0, "qc": 0, "sec": 0, "ext": 0,
                }
            stock = info.get("stock", 0)
            pred = info.get("predespacho", 0)
            disp = info.get("disponible", max(0, stock - pred))
            e["skus"].add(sku)
            e["stock"] += stock
            e["pred"] += pred
            if rol == "PRINCIPAL":
                e["ves"] += disp
            elif is_qc:
                e["qc"] += stock
            elif rol == "SECUNDARIO":
                e["sec"] += disp
            elif rol == "EXTERNO":
                e["ext"] += disp
    for e in aggr.values():
        sm = e["stock_minimo"]
        sv = e["ves"]
        if sm > 0 and sv < sm:
            e["salud"] = "CRITICO"
        elif sm > 0 and sv < sm * 2:
            e["salud"] = "ALERTA"
        else:
            e["salud"] = "BUENO"
    return aggr


def _line_qc_stock(sku: str, raw: dict, alm_config: dict, scope: str) -> int:
    """Stock del SKU en almacenes de control de calidad dentro del alcance."""
    total = 0
    for cod in _qc_codes(alm_config):
        if cod in raw and (scope != "control" or alm_config[cod].get("participa_control")):
            info = raw[cod].get(sku)
            if info:
                total += info.get("stock", 0)
    return total


def _write_role_resumen(ws, role_lines: dict, header_fill, white_font, red_fill, yellow_fill, green_fill):
    headers = ["Línea", "SKUs", "Disponible", "Estado"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    lines = sorted(role_lines.items(), key=lambda kv: min(_sku_info(s)["indice"] for s in kv[1]["skus"]))

    for linea, e in lines:
        disp_total = e["ves"] + e["sec"] + e["ext"]
        salud = e["salud"]
        ws.append([e["nombre"], len(e["skus"]), disp_total, salud])
        r = ws.max_row
        estado_cell = ws.cell(row=r, column=4)
        if salud == "CRITICO":
            estado_cell.fill = red_fill
        elif salud == "ALERTA":
            estado_cell.fill = yellow_fill
        else:
            estado_cell.fill = green_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12


def _write_classic_resumen(ws, grouped_data: dict, header_fill, white_font, red_fill, yellow_fill, green_fill):
    ws.append(["Línea", "SKUs", "Disponible", "Estado"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
    for linea, items in grouped_data.items():
        st = pre = disp = 0
        sku_count = len(items)
        for d in items:
            f = _row_fields(d)
            if f is None:
                continue
            st += f["st"]
            pre += f["pre"]
            disp += f["disp"]
        sm = 0
        for it in items:
            rf = _row_fields(it)
            if rf:
                cat = _sku_info(rf["sku"])
                sm = cat.get("stock_minimo", 0) or 0
                break
        if sm > 0 and disp < sm:
            estado = "CRITICO"
        elif sm > 0 and disp < sm * 2:
            estado = "ALERTA"
        else:
            estado = "BUENO"
        ws.append([linea, sku_count, disp, estado])
        r = ws.max_row
        est_cell = ws.cell(row=r, column=4)
        if estado == "CRITICO":
            est_cell.fill = red_fill
        elif estado == "ALERTA":
            est_cell.fill = yellow_fill
        else:
            est_cell.fill = green_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12


_AGG_HEADERS = {
    "Almacenes": ["Código", "Stock", "Pred.", "Disp.", "Alertas", "Críticos"],
    "SKUs por Categoría": ["Categoría", "SKUs", "Disp."],
    "Disponible por Línea": ["Línea", "Disponible", "Stock", "SKUs"],
    "Predespacho por Línea": ["Línea", "Pred.", "Stock", "SKUs"],
    "Productos en Alerta": ["SKU", "Descripción", "Disp.", "Almacén"],
    "SKUs sin Catálogo": ["SKU", "Descripción", "Disp."],
    "Productos Críticos": ["SKU", "Descripción", "Disp.", "Mínimo"],
    "SKUs alto predespacho": ["SKU", "Descripción", "Disp.", "Ratio"],
    "Transferencias Sugeridas": ["SKU", "Descripción", "Principal", "Secundario", "Sugerencia"],
    "SKUs por Almacén": ["SKU", "Descripción", "Almacén", "Disp."],
}


def _write_generic_table(ws, data_items: list, title: str, header_fill, white_font):
    """Exporta tablas agregadas (Almacenes, por Línea, Categorías, Predespacho) tal cual."""
    rows = [d for d in data_items if isinstance(d, (list, tuple))]
    if not rows:
        return
    key = next((k for k in _AGG_HEADERS if title.startswith(k)), None)
    headers = _AGG_HEADERS.get(key) if key else [f"Col {i + 1}" for i in range(max(len(r) for r in rows))]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
    for d in rows:
        vals = []
        for v in d[:len(headers)]:
            if isinstance(v, (list, tuple, set)):
                vals.append(", ".join(str(x) for x in v))
            else:
                vals.append(v)
        ws.append(vals)
    ws.column_dimensions['A'].width = 35
    for col in "BCD":
        ws.column_dimensions[col].width = 16


def _group_data_by_linea(data_items: list, real_skus: set) -> dict[str, list]:
    grouped_data = {}
    for item in data_items:
        fields = _row_fields(item)
        if fields is None:
            continue
        sku = fields["sku"]
        if real_skus and sku not in real_skus:
            continue
        linea = _sku_info(sku).get("linea", "")
        grouped_data.setdefault(linea, []).append(item)
    return grouped_data


def _write_linea_sheet(ws, items: list, role_line: dict | None, include_details: bool,
                        header_fill, white_font, red_fill, yellow_fill, green_fill,
                        raw: dict | None, alm_config: dict | None, scope: str):
    headers = ["SKU", "Descripción", "Categoría", "Stock", "Disponible"]
    if include_details:
        headers.append("Predespacho")
    headers.append("Estado")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    for d in items:
        f = _row_fields(d)
        if f is None:
            continue
        cat = _sku_info(f["sku"])
        disp = f["disp"]
        sm = cat.get("stock_minimo", 0) or 0
        if sm > 0 and disp < sm:
            estado = "CRITICO"
        elif sm > 0 and disp < sm * 2:
            estado = "ALERTA"
        else:
            estado = "BUENO"
        row_data = [f["sku"], f["desc"], cat.get("categoria", ""), f["st"], disp]
        if include_details:
            row_data.append(f["pre"])
        row_data.append(estado)
        ws.append(row_data)
        est_cell = ws.cell(row=ws.max_row, column=len(headers))
        if estado == "CRITICO":
            est_cell.fill = red_fill
        elif estado == "ALERTA":
            est_cell.fill = yellow_fill
        else:
            est_cell.fill = green_fill
        ws.cell(row=ws.max_row, column=len(headers)).alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    if include_details:
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
    else:
        ws.column_dimensions['F'].width = 12


def _write_excel_resumen(ws_resumen, role_lines, grouped_data, sorted_lineas, title,
                          include_summary, header_fill, white_font, red_fill, yellow_fill, green_fill):
    if role_lines:
        ws_resumen.title = "Resumen"
        _write_role_resumen(ws_resumen, role_lines, header_fill, white_font, red_fill, yellow_fill, green_fill)
    elif sorted_lineas:
        ws_resumen.title = "Resumen"
        _write_classic_resumen(ws_resumen, grouped_data, header_fill, white_font, red_fill, yellow_fill, green_fill)
    else:
        ws_resumen = None
    return ws_resumen


def export_to_excel(data_items: list, file_path: str, title: str, include_details: bool = False,
                    include_summary: bool = True, scope: str = "control",
                    raw: dict | None = None, alm_config: dict | None = None,
                    lineas_config: list | None = None):
    """
    Genera un archivo Excel profesional con hojas por línea y formato condicional.
    scope: "control" (VES + secundarios + QC) o "todos" (incluye externos).
    Con `raw`/`alm_config` presentes, el Resumen desglosa por rol (VES/QC/Sec./Ext.) y salud.
    """
    wb = Workbook()
    wb.properties.creator = APP_AUTHOR
    wb.properties.description = f"Reporte de stock generado por {APP_NAME} — {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    green_fill = PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid")

    skus = _extract_report_skus(data_items, raw) if raw else []
    real_skus = set(skus)
    role_lines = _build_role_report(skus, raw, alm_config, lineas_config, scope) if (skus and alm_config) else None

    grouped_data = _group_data_by_linea(data_items, real_skus)

    def line_indice(linea):
        for it in grouped_data[linea]:
            f = _row_fields(it)
            if f:
                return _sku_info(f["sku"])["indice"]
        return 9999

    sorted_lineas = sorted(grouped_data.keys(), key=line_indice)

    ws_resumen = None
    if include_summary:
        ws_resumen = wb.active
        ws_resumen = _write_excel_resumen(ws_resumen, role_lines, grouped_data, sorted_lineas, title,
                                          include_summary, header_fill, white_font, red_fill, yellow_fill, green_fill)

    for linea in sorted_lineas:
        ws = wb.create_sheet(title=linea[:31])
        rline = role_lines.get(linea) if role_lines else None
        _write_linea_sheet(ws, grouped_data[linea], rline, include_details,
                           header_fill, white_font, red_fill, yellow_fill, green_fill,
                           raw, alm_config, scope)

    if not sorted_lineas and not role_lines:
        if ws_resumen is None:
            ws_resumen = wb.active
            ws_resumen.title = (title[:31] or "Datos")
        _write_generic_table(ws_resumen, data_items, title, header_fill, white_font)
    else:
        if ws_resumen is None and "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        if wb.sheetnames:
            wb.active = 0

    wb.save(file_path)


CATALOG_COLUMNS = [
    ("orden", "Orden"),
    ("sku", "SKU"),
    ("nombre", "Nombre"),
    ("nombre_corto", "Nombre Corto"),
    ("linea", "Línea"),
    ("grupo", "Grupo"),
    ("tipo", "Tipo"),
    ("familia", "Familia"),
    ("categoria", "Categoría"),
    ("estado_linea", "Estado Línea"),
    ("un_bx", "Unidades/Caja"),
    ("peso_kg", "Peso (kg)"),
    ("precio", "Precio"),
    ("ean13", "EAN13"),
    ("ean14", "EAN14"),
    ("keywords", "Keywords"),
]

SIN_CATALOGO_COLUMNS = [
    ("sku", "SKU"),
    ("descripcion", "Descripción"),
    ("um", "UM"),
    ("linea", "Línea"),
    ("grupo", "Grupo"),
    ("tipo", "Tipo"),
    ("familia", "Familia"),
    ("categoria", "Categoría"),
    ("estado_linea", "Estado Línea"),
]


def _fill_excel_sheet(ws, columns: list, rows: list) -> None:
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    for col, (_, header) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    for r, item in enumerate(rows, start=2):
        for col, (key, _) in enumerate(columns, start=1):
            val = item.get(key)
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            ws.cell(row=r, column=col, value=val)

    for col, (key, _) in enumerate(columns, start=1):
        max_len = len(key)
        for r in range(2, min(len(rows) + 2, 400)):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"


def export_catalogo_to_excel(items: list, file_path: str, sin_catalogo_items: list | None = None) -> dict:
    """Genera un XLSX con el catálogo maestro (sin almacenes ni stock).

    Si se pasa `sin_catalogo_items` (SKUs de stock sin datos de catálogo),
    agrega una hoja "Sin Catálogo". Devuelve {catalogo, sin_catalogo} con los conteos.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"
    wb.properties.creator = APP_AUTHOR
    wb.properties.description = f"Catálogo maestro generado por {APP_NAME} — {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    _fill_excel_sheet(ws, CATALOG_COLUMNS, items)

    sin_items = sin_catalogo_items or []
    if sin_items:
        ws_sin = wb.create_sheet(title="Sin Catálogo")
        _fill_excel_sheet(ws_sin, SIN_CATALOGO_COLUMNS, sin_items)

    wb.save(file_path)
    return {"catalogo": len(items), "sin_catalogo": len(sin_items)}
